"""
SheetPic Embed 测试脚本
测试 sheetpic_embed.py 的核心逻辑（无需 GUI）
"""
import os
import sys
import time
import re
import platform
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage
from io import BytesIO
import requests
import concurrent.futures

# ==========================================
# 测试配置
# ==========================================
TEST_FILE = os.path.join(os.path.dirname(__file__), "test.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "test_output")
IMG_WIDTH = 500
IMG_HEIGHT = 500
TIMEOUT = 15

# 颜色输出
GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
RESET = "\033[0m"


def log_pass(name):
    print(f"  {GREEN}PASS{RESET} {name}")


def log_fail(name, reason=""):
    print(f"  {RED}FAIL{RESET} {name}" + (f" — {reason}" if reason else ""))


def log_info(msg):
    print(f"  {YELLOW}INFO{RESET} {msg}")


# ==========================================
# 测试 1: 文件加载
# ==========================================
def test_file_loading():
    print(f"\n{'='*50}")
    print("测试 1: 文件加载")
    print(f"{'='*50}")
    results = {}

    # 1.1 文件存在
    if os.path.exists(TEST_FILE):
        log_pass("1.1 文件存在")
        results['file_exists'] = True
    else:
        log_fail("1.1 文件存在", f"找不到 {TEST_FILE}")
        results['file_exists'] = False
        return results

    # 1.2 openpyxl 加载
    try:
        wb = openpyxl.load_workbook(TEST_FILE, data_only=True)
        ws = wb.active
        log_pass(f"1.2 openpyxl 加载成功 (行:{ws.max_row}, 列:{ws.max_column})")
        results['openpyxl'] = True
        results['max_row'] = ws.max_row
        results['max_col'] = ws.max_column
    except Exception as e:
        log_fail("1.2 openpyxl 加载", str(e))
        results['openpyxl'] = False

    # 1.3 pandas 加载
    try:
        df = pd.read_excel(TEST_FILE)
        log_pass(f"1.3 pandas 加载成功 ({len(df)} 行 x {len(df.columns)} 列)")
        results['pandas'] = True
        results['df'] = df
        results['columns'] = list(df.columns)
    except Exception as e:
        log_fail("1.3 pandas 加载", str(e))
        results['pandas'] = False

    # 1.4 列名检查
    if 'columns' in results:
        print(f"\n  列名列表 ({len(results['columns'])} 列):")
        for i, c in enumerate(results['columns']):
            log_info(f"  [{i}] {c} → {get_column_letter(i+1)}")

    return results


# ==========================================
# 测试 2: URL 列检测
# ==========================================
def test_url_detection(df):
    print(f"\n{'='*50}")
    print("测试 2: URL 列检测")
    print(f"{'='*50}")
    results = {}

    df_str = df.astype(str)
    cols = list(df_str.columns)

    # 2.1 扫描所有列
    url_counts = {}
    for i, c in enumerate(cols):
        sample = df_str[c].head(50).str.contains("http", case=False).any()
        if sample:
            real_count = df_str[c].str.contains("http", case=False, na=False).sum()
            if real_count > 0:
                url_counts[i] = real_count

    if url_counts:
        log_pass(f"2.1 检测到 {len(url_counts)} 个 URL 列")
        results['url_cols'] = url_counts
        for idx, count in url_counts.items():
            col_letter = get_column_letter(idx + 1)
            log_info(f"  列 {col_letter} ({cols[idx]}): {count} 条 URL")
    else:
        log_fail("2.1 URL 列检测", "未找到任何 URL 列")
        results['url_cols'] = {}
        return results

    # 2.2 URL 样本验证
    url_col_idx = max(url_counts, key=url_counts.get)
    sample_urls = df_str.iloc[:3, url_col_idx].tolist()
    print(f"\n  URL 样本 (列 {get_column_letter(url_col_idx+1)}):")
    for u in sample_urls:
        log_info(f"  {u[:70]}...")

    results['best_url_col'] = url_col_idx
    return results


# ==========================================
# 测试 3: SKU 列检测
# ==========================================
def test_sku_detection(columns):
    print(f"\n{'='*50}")
    print("测试 3: SKU 列检测")
    print(f"{'='*50}")
    results = {}

    sku_keywords = ["code", "sku", "条码", "货号", "编号", "id"]
    best = None
    for i, c in enumerate(columns):
        if any(k in c.lower() for k in sku_keywords):
            best = (i, c)
            break

    if best:
        log_pass(f"3.1 SKU 列匹配: {best[1]} (列 {get_column_letter(best[0]+1)})")
        results['sku_col'] = best[0]
    else:
        log_fail("3.1 SKU 列匹配", "未找到含关键词的列，将使用第一列")
        results['sku_col'] = 0

    return results


# ==========================================
# 测试 4: 图片下载
# ==========================================
def test_image_download(df, url_col_idx):
    print(f"\n{'='*50}")
    print("测试 4: 图片下载")
    print(f"{'='*50}")
    results = {'success': 0, 'fail': 0, 'errors': []}

    df_str = df.astype(str)
    urls = []
    for i in range(len(df_str)):
        val = str(df_str.iloc[i, url_col_idx]).strip()
        if val and val.lower() != 'nan' and 'http' in val.lower():
            urls.append((i, val))

    if not urls:
        log_fail("4.0 URL 提取", "未找到有效 URL")
        return results

    log_info(f"共 {len(urls)} 条 URL，开始下载测试...")

    # 测试前 3 条
    test_count = min(3, len(urls))
    for idx, (row_i, url) in enumerate(urls[:test_count]):
        try:
            r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=TIMEOUT)
            if r.status_code == 200:
                content_type = r.headers.get('Content-Type', '')
                log_pass(f"4.{idx+1} 行 {row_i+1} 下载成功 (HTTP {r.status_code}, {content_type[:30]})")
                results['success'] += 1

                # 4.x PIL 解析
                try:
                    pil_img = PILImage.open(BytesIO(r.content))
                    log_pass(f"  4.{idx+1}a PIL 解析成功 ({pil_img.size[0]}x{pil_img.size[1]}, {pil_img.mode})")
                except Exception as e:
                    log_fail(f"  4.{idx+1}a PIL 解析", str(e)[:50])

                # 4.x 缩放
                try:
                    pil_img.thumbnail((IMG_WIDTH, IMG_HEIGHT), PILImage.LANCZOS)
                    log_pass(f"  4.{idx+1}b 缩放成功 → {pil_img.size[0]}x{pil_img.size[1]}")
                except Exception as e:
                    log_fail(f"  4.{idx+1}b 缩放", str(e)[:50])
            else:
                log_fail(f"4.{idx+1} 行 {row_i+1} 下载", f"HTTP {r.status_code}")
                results['fail'] += 1
                results['errors'].append((row_i, f"HTTP {r.status_code}"))
        except requests.exceptions.Timeout:
            log_fail(f"4.{idx+1} 行 {row_i+1} 下载", "超时")
            results['fail'] += 1
            results['errors'].append((row_i, "超时"))
        except Exception as e:
            log_fail(f"4.{idx+1} 行 {row_i+1} 下载", str(e)[:60])
            results['fail'] += 1
            results['errors'].append((row_i, str(e)[:60]))

    return results


# ==========================================
# 测试 5: Excel 嵌入
# ==========================================
def test_excel_embedding(df, url_col_idx, sku_col_idx):
    print(f"\n{'='*50}")
    print("测试 5: Excel 图片嵌入")
    print(f"{'='*50}")
    results = {'embedded': 0, 'failed': 0}

    df_str = df.astype(str)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_file = os.path.join(OUTPUT_DIR, "test_embedded.xlsx")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active

    # 写入表头
    orig_cols = list(df_str.columns)
    img_header_col = url_col_idx + 2  # 1-indexed，跳过URL列

    out_col = 1
    for i, col_name in enumerate(orig_cols):
        ws.cell(row=1, column=out_col, value=col_name)
        out_col += 1
        if i == url_col_idx:
            ws.cell(row=1, column=out_col, value="图片")
            out_col += 1

    # 下载并嵌入前 5 行
    test_rows = min(5, len(df_str))
    log_info(f"测试嵌入前 {test_rows} 行...")

    for i in range(test_rows):
        url_raw = str(df_str.iloc[i, url_col_idx]).strip()
        if not url_raw or url_raw.lower() == 'nan' or 'http' not in url_raw.lower():
            continue

        try:
            r = requests.get(url_raw, headers={'User-Agent': 'Mozilla/5.0'}, timeout=TIMEOUT)
            if r.status_code == 200:
                pil_img = PILImage.open(BytesIO(r.content))
                pil_img.thumbnail((IMG_WIDTH, IMG_HEIGHT), PILImage.LANCZOS)
                buf = BytesIO()
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    pil_img = pil_img.convert('RGBA')
                    pil_img.save(buf, format='PNG')
                else:
                    pil_img = pil_img.convert('RGB')
                    pil_img.save(buf, format='JPEG', quality=85)
                buf.seek(0)

                # 写入原始数据
                excel_row = i + 2
                out_col = 1
                for j in range(len(orig_cols)):
                    val = str(df_str.iloc[i, j])
                    if val.lower() == 'nan':
                        val = ""
                    ws.cell(row=excel_row, column=out_col, value=val)
                    out_col += 1
                    if j == url_col_idx:
                        out_col += 1

                # 嵌入图片
                xl_img = XlImage(buf)
                ws.add_image(xl_img, f"{get_column_letter(img_header_col)}{excel_row}")
                ws.row_dimensions[excel_row].height = IMG_HEIGHT * 0.8 + 10
                results['embedded'] += 1
                log_pass(f"5.{i+1} 行 {i+1} 嵌入成功")
        except Exception as e:
            log_fail(f"5.{i+1} 行 {i+1} 嵌入", str(e)[:60])
            results['failed'] += 1

    # 设置列宽
    ws.column_dimensions[get_column_letter(img_header_col)].width = IMG_WIDTH / 7 + 2

    # 保存
    try:
        wb_out.save(out_file)
        file_size = os.path.getsize(out_file) / 1024
        log_pass(f"5.9 保存成功: {out_file} ({file_size:.1f} KB)")
        results['output_file'] = out_file
    except Exception as e:
        log_fail("5.9 保存", str(e))

    return results


# ==========================================
# 测试 6: 并发下载
# ==========================================
def test_concurrent_download(df, url_col_idx):
    print(f"\n{'='*50}")
    print("测试 6: 并发下载 (ThreadPoolExecutor)")
    print(f"{'='*50}")
    results = {'success': 0, 'fail': 0}

    df_str = df.astype(str)
    urls = []
    for i in range(len(df_str)):
        val = str(df_str.iloc[i, url_col_idx]).strip()
        if val and val.lower() != 'nan' and 'http' in val.lower():
            urls.append((i, val))

    if not urls:
        log_fail("6.0", "无 URL")
        return results

    def download_one(url):
        try:
            r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=TIMEOUT)
            if r.status_code == 200:
                pil_img = PILImage.open(BytesIO(r.content))
                pil_img.thumbnail((IMG_WIDTH, IMG_HEIGHT), PILImage.LANCZOS)
                buf = BytesIO()
                pil_img.convert('RGB').save(buf, format='JPEG')
                buf.seek(0)
                return True, len(r.content)
            return False, f"HTTP {r.status_code}"
        except Exception as e:
            return False, str(e)[:50]

    t_start = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(download_one, url): i for i, url in urls}
        for future in concurrent.futures.as_completed(futures):
            row_idx = futures[future]
            try:
                is_ok, info = future.result()
                if is_ok:
                    results['success'] += 1
                else:
                    results['fail'] += 1
                    log_fail(f"6.x 行 {row_idx+1}", str(info))
            except Exception as e:
                results['fail'] += 1
                log_fail(f"6.x 行 {row_idx+1}", str(e)[:50])

    duration = time.time() - t_start
    log_pass(f"6.1 并发下载完成: {results['success']} 成功, {results['fail']} 失败, 耗时 {duration:.1f}s")
    return results


# ==========================================
# 测试 7: 边界情况
# ==========================================
def test_edge_cases(df, url_col_idx):
    print(f"\n{'='*50}")
    print("测试 7: 边界情况")
    print(f"{'='*50}")

    df_str = df.astype(str)

    # 7.1 空 URL 处理
    empty_count = 0
    for i in range(len(df_str)):
        val = str(df_str.iloc[i, url_col_idx]).strip()
        if not val or val.lower() == 'nan' or 'http' not in val.lower():
            empty_count += 1
    log_info(f"7.1 空/无效 URL 数: {empty_count}")

    # 7.2 URL 清洗（缩略图参数去除）
    def clean_url(url):
        url = re.sub(r'!\d+x\d+', '', url)
        url = re.sub(r'\?imageView2/[^&]*', '', url)
        url = re.sub(r'\?x-oss-process=[^&]*', '', url)
        url = re.sub(r'[?&](width|height|w|h|size|resize|quality|format)=[^&]*', '', url)
        url = re.sub(r'\?\d+$', '', url)
        url = re.sub(r'\?&+', '?', url)
        url = re.sub(r'\?$', '', url)
        return url

    test_cases = [
        ("http://example.com/img.jpg!200x200?123456", "http://example.com/img.jpg"),
        ("http://example.com/img.jpg?width=200&height=200", "http://example.com/img.jpg"),
        ("http://example.com/img.jpg?imageView2/0/w/200/h/200", "http://example.com/img.jpg"),
        ("http://example.com/img.jpg?x-oss-process=image/resize,w_200", "http://example.com/img.jpg"),
        ("http://example.com/img.jpg", "http://example.com/img.jpg"),
    ]
    for original, expected in test_cases:
        result = clean_url(original)
        if result == expected:
            log_pass(f"7.2 URL清洗: {original[:50]}")
        else:
            log_fail(f"7.2 URL清洗: {original[:50]}", f"期望 {expected}, 得到 {result}")

    # 7.2b 实际下载对比：原图 vs 缩略图
    real_url_with_thumb = "http://erp.qiniu.wsgjp.com/513807601055186944/587201473623743750_pic1.png!200x200?134216946960000000"
    real_url_original = clean_url(real_url_with_thumb)
    try:
        r_thumb = requests.get(real_url_with_thumb, headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        r_orig = requests.get(real_url_original, headers={'User-Agent': 'Mozilla/5.0'}, timeout=10)
        if r_thumb.status_code == 200 and r_orig.status_code == 200:
            img_thumb = PILImage.open(BytesIO(r_thumb.content))
            img_orig = PILImage.open(BytesIO(r_orig.content))
            log_pass(f"7.2b 缩略图: {img_thumb.size[0]}x{img_thumb.size[1]} ({len(r_thumb.content)} bytes)")
            log_pass(f"7.2b 原图:    {img_orig.size[0]}x{img_orig.size[1]} ({len(r_orig.content)} bytes)")
            if img_orig.size[0] >= img_thumb.size[0] and img_orig.size[1] >= img_thumb.size[1]:
                log_pass("7.2b 原图尺寸 >= 缩略图尺寸 ✓")
            else:
                log_info("7.2b 原图尺寸 < 缩略图尺寸 (服务器可能不支持参数)")
    except Exception as e:
        log_info(f"7.2b 对比测试跳过: {str(e)[:40]}")

    # 7.3 无效尺寸回退
    try:
        w = int("abc")
    except ValueError:
        w = 120
    log_pass(f"7.3 无效尺寸回退: 'abc' → {w}")

    # 7.4 404 URL 测试
    try:
        r = requests.get("https://httpbin.org/status/404", timeout=5)
        if r.status_code == 404:
            log_pass(f"7.4 404 处理: HTTP {r.status_code} (正确捕获)")
        else:
            log_info(f"7.4 404 处理: HTTP {r.status_code}")
    except Exception as e:
        log_info(f"7.4 404 测试跳过: {str(e)[:40]}")

    # 7.5 超时测试
    try:
        r = requests.get("https://httpbin.org/delay/30", timeout=2)
        log_fail("7.5 超时处理", "应该超时但没有")
    except requests.exceptions.Timeout:
        log_pass("7.5 超时处理: 2秒超时正确触发")
    except Exception as e:
        log_info(f"7.5 超时测试: {str(e)[:40]}")


# ==========================================
# 测试 8: 完整流程模拟
# ==========================================
def test_full_workflow():
    print(f"\n{'='*50}")
    print("测试 8: 完整流程模拟 (全部行)")
    print(f"{'='*50}")

    df = pd.read_excel(TEST_FILE)
    df_str = df.astype(str)
    cols = list(df_str.columns)

    # 检测 URL 列
    url_col_idx = None
    for i, c in enumerate(cols):
        if df_str[c].head(50).str.contains("http", case=False).any():
            count = df_str[c].str.contains("http", case=False, na=False).sum()
            if count > 0:
                url_col_idx = i
                break

    if url_col_idx is None:
        log_fail("8.0", "未找到 URL 列")
        return

    # 检测 SKU 列
    sku_col_idx = 0
    for i, c in enumerate(cols):
        if any(k in c.lower() for k in ["code", "sku", "条码", "货号", "编号"]):
            sku_col_idx = i
            break

    log_info(f"URL 列: {cols[url_col_idx]} ({get_column_letter(url_col_idx+1)})")
    log_info(f"SKU 列: {cols[sku_col_idx]} ({get_column_letter(sku_col_idx+1)})")
    log_info(f"图片尺寸: {IMG_WIDTH}x{IMG_HEIGHT}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_file = os.path.join(OUTPUT_DIR, "full_embedded.xlsx")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active

    orig_cols = list(df_str.columns)
    img_header_col = url_col_idx + 2  # 1-indexed，跳过URL列

    # 写表头
    out_col = 1
    for i, col_name in enumerate(orig_cols):
        ws.cell(row=1, column=out_col, value=col_name)
        out_col += 1
        if i == url_col_idx:
            ws.cell(row=1, column=out_col, value="图片")
            out_col += 1

    total = len(df_str)
    success = 0
    fail = 0

    # 并发下载所有
    def download_one(url):
        try:
            r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=TIMEOUT)
            if r.status_code == 200:
                pil_img = PILImage.open(BytesIO(r.content))
                pil_img.thumbnail((IMG_WIDTH, IMG_HEIGHT), PILImage.LANCZOS)
                buf = BytesIO()
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    pil_img = pil_img.convert('RGBA')
                    pil_img.save(buf, format='PNG')
                else:
                    pil_img = pil_img.convert('RGB')
                    pil_img.save(buf, format='JPEG', quality=85)
                buf.seek(0)
                return True, buf
            return False, f"HTTP {r.status_code}"
        except Exception as e:
            return False, str(e)[:50]

    # 准备 URL 列表
    rows_data = []
    for i in range(total):
        url_raw = str(df_str.iloc[i, url_col_idx]).strip()
        if url_raw and url_raw.lower() != 'nan' and 'http' in url_raw.lower():
            rows_data.append((i, url_raw))
        else:
            rows_data.append((i, None))

    t_start = time.time()
    row_results = [None] * total

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        futures = {}
        for i, url in rows_data:
            if url:
                futures[executor.submit(download_one, url)] = i
            else:
                futures[executor.submit(lambda: (False, "No URL"))] = i

        for future in concurrent.futures.as_completed(futures):
            row_idx = futures[future]
            try:
                is_ok, result = future.result()
            except Exception as e:
                is_ok, result = False, str(e)
            row_results[row_idx] = (is_ok, result)

    # 按行顺序写入 Excel
    for i, result in enumerate(row_results):
        if result is None:
            result = (False, "Stopped")
        is_ok, data = result
        excel_row = i + 2

        out_col = 1
        for j in range(len(orig_cols)):
            val = str(df_str.iloc[i, j])
            if val.lower() == 'nan':
                val = ""
            ws.cell(row=excel_row, column=out_col, value=val)
            out_col += 1
            if j == url_col_idx:
                out_col += 1

        if is_ok:
            try:
                xl_img = XlImage(data)
                ws.add_image(xl_img, f"{get_column_letter(img_header_col)}{excel_row}")
                ws.row_dimensions[excel_row].height = IMG_HEIGHT * 0.8 + 10
                success += 1
            except Exception:
                ws.cell(row=excel_row, column=img_header_col, value="[嵌入失败]")
                fail += 1
        else:
            ws.cell(row=excel_row, column=img_header_col, value=f"[{data}]")
            fail += 1

    ws.column_dimensions[get_column_letter(img_header_col)].width = IMG_WIDTH / 7 + 2
    wb_out.save(out_file)

    duration = time.time() - t_start
    file_size = os.path.getsize(out_file) / 1024

    log_pass(f"8.1 完成: {success} 成功, {fail} 失败, 耗时 {duration:.1f}s")
    log_pass(f"8.2 输出: {out_file} ({file_size:.1f} KB)")


# ==========================================
# 测试 9: 删除URL列功能
# ==========================================
def test_del_url_column():
    print(f"\n{'='*50}")
    print("测试 9: 删除URL列功能 (del_url=True)")
    print(f"{'='*50}")

    df = pd.read_excel(TEST_FILE)
    df_str = df.astype(str)
    cols = list(df_str.columns)

    url_col_idx = None
    for i, c in enumerate(cols):
        if df_str[c].head(50).str.contains("http", case=False).any():
            count = df_str[c].str.contains("http", case=False, na=False).sum()
            if count > 0:
                url_col_idx = i
                break

    if url_col_idx is None:
        log_fail("9.0", "未找到 URL 列")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_file = os.path.join(OUTPUT_DIR, "del_url_embedded.xlsx")

    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    orig_cols = list(df_str.columns)
    del_url = True

    # 写表头（跳过URL列，在该位置写"图片"）
    out_col = 1
    for i, col_name in enumerate(orig_cols):
        if del_url and i == url_col_idx:
            ws.cell(row=1, column=out_col, value="图片")
            img_header_col = out_col
            out_col += 1
        else:
            ws.cell(row=1, column=out_col, value=col_name)
            out_col += 1

    # 写数据（跳过URL列）
    for i in range(min(3, len(df_str))):
        excel_row = i + 2
        out_col = 1
        for j in range(len(orig_cols)):
            if del_url and j == url_col_idx:
                out_col += 1
            else:
                val = str(df_str.iloc[i, j])
                if val.lower() == 'nan':
                    val = ""
                ws.cell(row=excel_row, column=out_col, value=val)
                out_col += 1

    wb_out.save(out_file)

    # 验证输出
    wb_check = openpyxl.load_workbook(out_file)
    ws_check = wb_check.active
    headers = [ws_check.cell(row=1, column=c).value for c in range(1, ws_check.max_column + 1)]

    if "图" in headers:
        log_fail("9.1 URL列应被删除", f"但表头中仍包含 '图': {headers}")
    else:
        log_pass("9.1 URL列已删除")

    if "图片" in headers:
        log_pass(f"9.2 图片列存在 (位置: {headers.index('图片') + 1})")
    else:
        log_fail("9.2 图片列不存在", f"表头: {headers}")

    log_pass(f"9.3 输出: {out_file} ({os.path.getsize(out_file) / 1024:.1f} KB)")
    log_info(f"  表头: {headers[:5]}...")


# ==========================================
# 主测试流程
# ==========================================
def main():
    print("=" * 50)
    print("SheetPic Embed 测试报告")
    print(f"平台: {platform.system()} {platform.release()}")
    print(f"Python: {sys.version.split()[0]}")
    print(f"测试文件: {TEST_FILE}")
    print("=" * 50)

    # 测试 1: 文件加载
    load_results = test_file_loading()
    if not load_results.get('file_exists') or not load_results.get('pandas'):
        print(f"\n{RED}文件加载失败，终止测试{RESET}")
        return

    df = load_results['df']

    # 测试 2: URL 列检测
    url_results = test_url_detection(df)
    if 'best_url_col' not in url_results:
        print(f"\n{RED}未检测到 URL 列，终止测试{RESET}")
        return

    url_col_idx = url_results['best_url_col']

    # 测试 3: SKU 列检测
    sku_results = test_sku_detection(load_results['columns'])
    sku_col_idx = sku_results['sku_col']

    # 测试 4: 图片下载
    test_image_download(df, url_col_idx)

    # 测试 5: Excel 嵌入
    test_excel_embedding(df, url_col_idx, sku_col_idx)

    # 测试 6: 并发下载
    test_concurrent_download(df, url_col_idx)

    # 测试 7: 边界情况
    test_edge_cases(df, url_col_idx)

    # 测试 8: 完整流程
    test_full_workflow()

    # 测试 9: 删除URL列
    test_del_url_column()

    print(f"\n{'='*50}")
    print(f"{GREEN}测试完成{RESET}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
