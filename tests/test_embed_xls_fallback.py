import os
import sys
from io import BytesIO

import openpyxl
import pandas as pd
from PIL import Image as PILImage

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)


def test_prepare_embed_image_bytes_flattens_transparency_to_white_jpg_with_border():
    from sheetpic import _prepare_embed_image_bytes

    src = PILImage.new('RGBA', (100, 50), (0, 0, 0, 0))
    red = PILImage.new('RGBA', (50, 30), (200, 0, 0, 255))
    src.paste(red, (25, 10))

    out = PILImage.open(_prepare_embed_image_bytes(src))

    assert out.format == 'JPEG'
    assert out.mode == 'RGB'
    assert out.size == (110, 56)
    assert min(out.getpixel((1, 1))) > 245
    assert min(out.getpixel((6, 4))) > 245
    center = out.getpixel((55, 28))
    assert center[0] > 150
    assert center[1] < 80
    assert center[2] < 80


def test_prepare_embed_image_bytes_can_preserve_transparency_as_png_with_border():
    from sheetpic import EMBED_BG_TRANSPARENT, _prepare_embed_image_bytes

    src = PILImage.new('RGBA', (100, 50), (0, 0, 0, 0))
    red = PILImage.new('RGBA', (50, 30), (200, 0, 0, 255))
    src.paste(red, (25, 10))

    out = PILImage.open(_prepare_embed_image_bytes(src, bg_mode=EMBED_BG_TRANSPARENT))

    assert out.format == 'PNG'
    assert out.mode == 'RGBA'
    assert out.size == (110, 56)
    assert out.getpixel((1, 1))[3] == 0
    assert out.getpixel((6, 4))[3] == 0
    center = out.getpixel((55, 28))
    assert center[:3] == (200, 0, 0)
    assert center[3] == 255


def test_transparent_mode_does_not_fake_alpha_for_opaque_sources():
    from sheetpic import EMBED_BG_TRANSPARENT, _prepare_embed_image_bytes

    src = PILImage.new('RGB', (100, 50), 'blue')
    out = PILImage.open(_prepare_embed_image_bytes(src, bg_mode=EMBED_BG_TRANSPARENT))

    assert out.format == 'JPEG'
    assert out.mode == 'RGB'
    assert out.size == (110, 56)
    assert min(out.getpixel((1, 1))) > 245


def test_prepare_embed_image_bytes_keeps_max_dim_including_border():
    from sheetpic import _prepare_embed_image_bytes

    src = PILImage.new('RGB', (1000, 800), 'blue')
    out = PILImage.open(_prepare_embed_image_bytes(src, max_dim=500))

    assert out.format == 'JPEG'
    assert max(out.size) <= 500


def test_write_original_xls_falls_back_to_new_xlsx(tmp_path):
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.file_path = str(tmp_path / 'legacy.xls')
    app.header_row = 2
    app.df = pd.DataFrame({
        '图片': ['http://x/1.jpg'],
        '条码': ['A001'],
    })
    logs = []

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def get(self):
            return str(tmp_path)

    with open(app.file_path, 'wb') as f:
        f.write(b'not an xlsx file')

    app.root = _RootStub()
    app.entry_dest = _EntryStub()
    app.log = logs.append

    out_file, ws, wb_out, img_header_col, header_row_excel = app._embed_setup_original('legacy', 0)

    try:
        assert out_file == str(tmp_path / 'legacy_Embedded.xlsx')
        assert img_header_col == 2
        assert header_row_excel == 3
        assert ws.cell(row=3, column=1).value == '图片'
        assert ws.cell(row=3, column=2).value == '图片'
        assert app._embed_setup_used_original is False
        assert any('.xls' in msg for msg in logs)
    finally:
        wb_out.close()


def test_xls_fallback_run_writes_source_rows(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.file_path = str(tmp_path / 'legacy.xls')
    app.header_row = 0
    app.df = pd.DataFrame({
        '图片': ['http://x/1.jpg'],
        '条码': ['A001'],
        '商品全名': ['测试商品'],
    })
    app.embed_url_col_idx = 0
    app.embed_sku_col_idx = 1
    app.is_running = True
    app._process_start_time = 0
    logs = []
    status = []

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _VarStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _WidgetStub:
        def config(self, **kwargs):
            status.append(kwargs)

    class _ProgressStub(dict):
        def stop(self):
            pass

    with open(app.file_path, 'wb') as f:
        f.write(b'not an xlsx file')

    img_bytes = BytesIO()
    PILImage.new('RGB', (10, 10), 'white').save(img_bytes, format='PNG')
    img_payload = img_bytes.getvalue()

    def _download(_url, _max_dim=None, _bg_mode=None):
        return True, BytesIO(img_payload)

    app.root = _RootStub()
    app.entry_dest = _EntryStub(str(tmp_path))
    app.entry_max_dim = _EntryStub('500')
    app.var_original = _VarStub(False)
    app.var_write_original = _VarStub(True)
    app.progress = _ProgressStub()
    app.lbl_status = _WidgetStub()
    app.btn_run = _WidgetStub()
    app.btn_stop = _WidgetStub()
    app.log = logs.append
    app.download_to_bytesio = _download
    app._open_folder = lambda _path: None
    monkeypatch.setattr(sheetpic.messagebox, 'showinfo', lambda *_args, **_kwargs: None)

    app.run_embed_process()

    out_file = tmp_path / 'legacy_Embedded.xlsx'
    assert out_file.exists()

    wb = openpyxl.load_workbook(out_file)
    try:
        ws = wb.active
        assert ws.cell(row=1, column=1).value == '图片'
        assert ws.cell(row=1, column=2).value == '条码'
        assert ws.cell(row=1, column=3).value == '图片'
        assert ws.cell(row=1, column=4).value == '商品全名'
        assert ws.cell(row=2, column=1).value == 'http://x/1.jpg'
        assert ws.cell(row=2, column=2).value == 'A001'
        assert ws.cell(row=2, column=4).value == '测试商品'
    finally:
        wb.close()


def test_embed_process_uses_url_library_without_url_column(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.file_path = "Clipboard"
    app.header_row = 0
    app.df = pd.DataFrame({
        '条码': ['A001', 'A002'],
        '商品全名': ['测试商品1', '测试商品2'],
    })
    app.embed_url_col_idx = None
    app.embed_sku_col_idx = 0
    app.embed_use_url_library = True
    app.url_library = {'A001': 'http://x/1.webp'}
    app.is_running = True
    app._process_start_time = 0
    logs = []
    status = []
    downloaded = []

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _VarStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _WidgetStub:
        def config(self, **kwargs):
            status.append(kwargs)

    class _ProgressStub(dict):
        def stop(self):
            pass

    img_bytes = BytesIO()
    PILImage.new('RGB', (10, 10), 'white').save(img_bytes, format='JPEG')
    img_payload = img_bytes.getvalue()

    def _download(url, _max_dim=None, _bg_mode=None):
        downloaded.append((url, _bg_mode))
        return True, BytesIO(img_payload)

    app.root = _RootStub()
    app.entry_dest = _EntryStub(str(tmp_path))
    app.entry_max_dim = _EntryStub('500')
    app.var_original = _VarStub(False)
    app.var_write_original = _VarStub(False)
    app.progress = _ProgressStub()
    app.lbl_status = _WidgetStub()
    app.btn_run = _WidgetStub()
    app.btn_stop = _WidgetStub()
    app.log = logs.append
    app.download_to_bytesio = _download
    app._open_folder = lambda _path: None
    monkeypatch.setattr(sheetpic.messagebox, 'showinfo', lambda *_args, **_kwargs: None)

    app.run_embed_process()

    out_file = tmp_path / 'Clipboard_Embedded.xlsx'
    assert out_file.exists()
    assert downloaded == [('http://x/1.webp', 'white')]
    assert any('URL库匹配: 1 / 2' in msg for msg in logs)

    wb = openpyxl.load_workbook(out_file)
    try:
        ws = wb.active
        assert ws.cell(row=1, column=1).value == '条码'
        assert ws.cell(row=1, column=2).value == '图片'
        assert ws.cell(row=1, column=3).value == '商品全名'
        assert ws.cell(row=2, column=1).value == 'A001'
        assert ws.cell(row=3, column=1).value == 'A002'
        assert ws.cell(row=3, column=2).value == app.T['msg_dl_fail']
    finally:
        wb.close()


def test_embed_process_writes_selected_url_library_fields(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.file_path = "Clipboard"
    app.header_row = 0
    app.df = pd.DataFrame({
        '条码': ['A001', 'A002'],
        '数量': [2, 3],
    })
    app.embed_url_col_idx = None
    app.embed_sku_col_idx = 0
    app.embed_use_url_library = True
    app.url_library = {'A001': 'http://x/1.webp'}
    app.url_library_records = {
        'A001': {'商品名称': '库商品1', '品牌': '库品牌'},
    }
    app.url_library_field_names = ['商品名称', '品牌']
    app.url_library_selected_fields = ['商品名称', '品牌']
    app.is_running = True
    app._process_start_time = 0
    logs = []
    status = []

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _VarStub:
        def __init__(self, value):
            self.value = value

        def get(self):
            return self.value

    class _WidgetStub:
        def config(self, **kwargs):
            status.append(kwargs)

    class _ProgressStub(dict):
        def stop(self):
            pass

    img_bytes = BytesIO()
    PILImage.new('RGB', (10, 10), 'white').save(img_bytes, format='JPEG')
    img_payload = img_bytes.getvalue()

    def _download(_url, _max_dim=None, _bg_mode=None):
        return True, BytesIO(img_payload)

    app.root = _RootStub()
    app.entry_dest = _EntryStub(str(tmp_path))
    app.entry_max_dim = _EntryStub('500')
    app.var_original = _VarStub(False)
    app.var_write_original = _VarStub(False)
    app.progress = _ProgressStub()
    app.lbl_status = _WidgetStub()
    app.btn_run = _WidgetStub()
    app.btn_stop = _WidgetStub()
    app.log = logs.append
    app.download_to_bytesio = _download
    app._open_folder = lambda _path: None
    monkeypatch.setattr(sheetpic.messagebox, 'showinfo', lambda *_args, **_kwargs: None)

    app.run_embed_process()

    out_file = tmp_path / 'Clipboard_Embedded.xlsx'
    wb = openpyxl.load_workbook(out_file)
    try:
        ws = wb.active
        assert ws.cell(row=1, column=1).value == '条码'
        assert ws.cell(row=1, column=2).value == '图片'
        assert ws.cell(row=1, column=3).value == '商品名称'
        assert ws.cell(row=1, column=4).value == '品牌'
        assert ws.cell(row=1, column=5).value == '数量'
        assert ws.cell(row=2, column=1).value == 'A001'
        assert ws.cell(row=2, column=3).value == '库商品1'
        assert ws.cell(row=2, column=4).value == '库品牌'
        assert ws.cell(row=2, column=5).value == '2'
        assert ws.cell(row=3, column=3).value in (None, '')
        assert ws.cell(row=3, column=4).value in (None, '')
    finally:
        wb.close()


def test_extract_download_retries_two_timeouts_then_succeeds(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.is_running = True
    calls = []

    class _Response:
        status_code = 200
        headers = {'Content-Type': 'image/jpeg', 'Content-Length': '3'}

        def iter_content(self, _chunk_size):
            yield b'abc'

    def _get(*_args, **_kwargs):
        calls.append(1)
        if len(calls) <= 2:
            raise sheetpic.requests.exceptions.Timeout()
        return _Response()

    monkeypatch.setattr(sheetpic.requests, 'get', _get)

    ok, msg = app.download_url('http://x/slow.jpg', 'slow', str(tmp_path))

    assert ok is True
    assert msg == 'OK'
    assert len(calls) == 3
    assert (tmp_path / 'slow.jpg').read_bytes() == b'abc'


def test_extract_retry_process_only_retries_failed_tasks(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.is_running = True
    app._process_start_time = 0
    app.extract_failed_tasks = []
    logs = []
    calls = []

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def get(self):
            return str(tmp_path)

    class _WidgetStub:
        def __init__(self):
            self.state = None
            self.text = None

        def config(self, **kwargs):
            if 'state' in kwargs:
                self.state = kwargs['state']
            if 'text' in kwargs:
                self.text = kwargs['text']

    class _ProgressStub(dict):
        def stop(self):
            pass

    retry_tasks = [
        {'url': 'http://x/ok.jpg', 'filename_base': 'ok', 'out_dir': str(tmp_path)},
        {'url': 'http://x/fail.jpg', 'filename_base': 'fail', 'out_dir': str(tmp_path)},
    ]

    def _download(_url, filename_base, _out_dir):
        calls.append(filename_base)
        if filename_base == 'ok':
            return True, 'OK'
        return False, 'still failed'

    app.root = _RootStub()
    app.entry_dest = _EntryStub()
    app.progress = _ProgressStub()
    app.lbl_status = _WidgetStub()
    app.btn_run = _WidgetStub()
    app.btn_stop = _WidgetStub()
    app.btn_retry = _WidgetStub()
    app.log = logs.append
    app.download_url = _download
    app._open_folder = lambda _path: None
    monkeypatch.setattr(sheetpic.messagebox, 'showinfo', lambda *_args, **_kwargs: None)

    app.run_extract_retry_process(retry_tasks)

    assert sorted(calls) == ['fail', 'ok']
    assert app.extract_failed_tasks == [retry_tasks[1]]
    assert app.btn_run.state == 'normal'
    assert app.btn_stop.state == 'disabled'
    assert app.btn_retry.state == 'normal'
    assert any('still failed' in msg for msg in logs)


def test_extract_retry_process_skips_existing_same_name(tmp_path, monkeypatch):
    import sheetpic
    from sheetpic import LANG_MAP, SheetPicApp

    app = SheetPicApp.__new__(SheetPicApp)
    app.T = LANG_MAP['zh']
    app.is_running = True
    app._process_start_time = 0
    app.extract_failed_tasks = []
    calls = []
    logs = []
    (tmp_path / 'ok.jpg').write_bytes(b'existing')

    class _RootStub:
        def after(self, _delay, fn, *args):
            fn(*args)

    class _EntryStub:
        def get(self):
            return str(tmp_path)

    class _WidgetStub:
        def __init__(self):
            self.state = None

        def config(self, **kwargs):
            self.state = kwargs.get('state', self.state)

    class _ProgressStub(dict):
        def stop(self):
            pass

    def _download(*_args):
        calls.append(_args)
        return True, 'OK'

    app.root = _RootStub()
    app.entry_dest = _EntryStub()
    app.progress = _ProgressStub()
    app.lbl_status = _WidgetStub()
    app.btn_run = _WidgetStub()
    app.btn_stop = _WidgetStub()
    app.btn_retry = _WidgetStub()
    app.log = logs.append
    app.download_url = _download
    app._open_folder = lambda _path: None
    monkeypatch.setattr(sheetpic.messagebox, 'showinfo', lambda *_args, **_kwargs: None)

    app.run_extract_retry_process([
        {'url': 'http://x/ok.jpg', 'filename_base': 'ok', 'out_dir': str(tmp_path)}
    ])

    assert calls == []
    assert app.extract_failed_tasks == []
    assert app.btn_retry.state == 'disabled'
    assert any('同名跳过' in msg for msg in logs)
