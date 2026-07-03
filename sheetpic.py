"""
SheetPic - 批量图片提取 & 嵌入 工具
支持两种模式:
  - 提取图片: 从Excel下载/导出嵌入图片
  - 嵌入图片: 将URL图片下载后嵌入Excel单元格
"""
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import math
import os
import threading
import platform
import concurrent.futures
from io import BytesIO
import webbrowser
import datetime
import mimetypes
import re
import json
import sys
import time
import subprocess
import urllib.request


class _LazyImport:
    """Import heavy optional modules only when a feature actually needs them."""

    def __init__(self, loader):
        self._loader = loader
        self._module = None

    def _load(self):
        if self._module is None:
            self._module = self._loader()
        return self._module

    def __getattr__(self, name):
        return getattr(self._load(), name)


def _load_pandas():
    import pandas as _pd
    return _pd


def _load_openpyxl():
    import openpyxl as _openpyxl
    return _openpyxl


def _load_requests():
    import requests as _requests
    return _requests


def _load_pil_image():
    from PIL import Image as _PILImage
    return _PILImage


pd = _LazyImport(_load_pandas)
openpyxl = _LazyImport(_load_openpyxl)
requests = _LazyImport(_load_requests)
PILImage = _LazyImport(_load_pil_image)


def get_column_letter(*args, **kwargs):
    from openpyxl.utils import get_column_letter as _get_column_letter
    return _get_column_letter(*args, **kwargs)


def column_index_from_string(*args, **kwargs):
    from openpyxl.utils import column_index_from_string as _column_index_from_string
    return _column_index_from_string(*args, **kwargs)


def XlImage(*args, **kwargs):
    from openpyxl.drawing.image import Image as _XlImage
    return _XlImage(*args, **kwargs)


def _resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def _set_window_icon(root):
    if platform.system() != "Windows":
        return
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(f"Andre.SheetPic.{APP_VERSION}")
    except Exception:
        pass

    icon_path = _resource_path("icon.ico")
    if not os.path.exists(icon_path):
        return
    try:
        root.iconbitmap(default=icon_path)
    except tk.TclError:
        try:
            root.iconbitmap(icon_path)
        except tk.TclError:
            pass

# ==========================================
# 版本号
# ==========================================
APP_VERSION = "1.0.44"

# ==========================================
# 语言与配置
# ==========================================
LANG_MAP = {
    'zh': {
        'title': "表图 - 表格图片提取 & 嵌入工具",
        'menu_lang': "语言",
        'footer_text': "Build by Andre  |  v{}",
        'tab_extract': "提取图片",
        'tab_embed': "嵌入图片",
        'sec_source': "数据来源",
        'sec_settings': "匹配与保存",
        'sec_embed_settings': "嵌入设置",
        'btn_browse': "📂 选择文件",
        'btn_clip': "📋 剪贴板",
        'lbl_dest': "保存位置:",
        'btn_dest': "修改",
        'lbl_sheet': "工作表:",
        'unnamed': "未命名",
        # Extract
        'lbl_img': "图片来源 (默认智能合并)",
        'lbl_code': "文件名列 (ID/SKU)",
        'opt_auto': "★ [智能合并] 优先下载数据最多的列 (推荐)",
        'type_url': "[链接] {} (含 {} 个URL)",
        'type_img': "[图片] {} (含 {} 张嵌入图)",
        'lbl_extract_bg': "提取背景",
        'opt_extract_bg_original': "保留原图",
        'opt_extract_bg_white': "白色背景",
        'lbl_extract_shape': "背景形状",
        'opt_extract_shape_original': "原比例",
        'opt_extract_shape_square': "正方形",
        'msg_skip': "❌ {}: [空] 未检测到有效图片",
        'done_msg': "耗时: {:.1f}s\n成功: {}\n失败: {}\n跳过: {}\n保存至: {}",
        # Embed
        'lbl_url_col': "图片URL列 (含链接的列)",
        'lbl_sku_col': "SKU/ID列 (用于排序)",
        'lbl_img_size': "最大边长 (px)",
        'lbl_url_library': "URL库: {} 条",
        'lbl_url_library_fields': "URL库字段: {} 个",
        'btn_import_url_lib': "导入URL库",
        'btn_clear_url_lib': "清空URL库",
        'btn_select_all_fields': "全选",
        'btn_clear_fields': "不选",
        'lbl_img_bg': "图片背景",
        'opt_bg_white': "白底 JPG",
        'opt_bg_transparent': "保留透明 PNG",
        'chk_original': "插入原图 (不缩放)",
        'chk_write_original': "写入原文件 (保留格式)",
        'msg_no_url': "❌ 未检测到包含URL的列",
        'opt_url_library': "[URL库] 按SKU/ID匹配 ({} 条)",
        'msg_url_lib_imported': "✅ URL库已导入: 新增/更新 {} 条，当前共 {} 条 ({})",
        'msg_url_lib_empty': "❌ URL库文件为空或未找到有效映射",
        'msg_url_lib_no_cols': "❌ URL库需要至少一列条码/SKU和一列图片URL",
        'msg_url_lib_cleared': "URL库已清空",
        'msg_url_lib_clear_confirm': "确定清空已保存的URL库吗？",
        'msg_use_url_library': "未检测到URL列，将使用URL库按SKU/ID匹配",
        'msg_url_lib_matches': "URL库匹配: {} / {}",
        'msg_embed_done': "耗时: {:.1f}s\n嵌入成功: {}\n下载失败: {}\n输出文件: {}",
        'msg_dl_fail': "[下载失败]",
        'msg_dl_skip': "[无URL]",
        'msg_same_name_skip': "⏭ {}: [同名跳过] 文件已存在",
        'msg_invalid_url': "⚠️ {}: [无效URL] {}",
        'msg_conn_err': "❌ {}: [连接失败] {}",
        'msg_ssl_err': "❌ {}: [SSL错误] {}",
        'msg_too_large': "❌ {}: [文件过大] {}MB",
        'msg_bad_image': "❌ {}: [图片格式错误] {}",
        'log_embed_start': "开始嵌入图片处理...",
        'log_embed_save': "正在保存Excel文件...",
        'log_embed_format_fallback': "{} 不支持保留格式写入，已自动改为新建 .xlsx 输出。",
        'msg_embed_error': "❌ 嵌入处理失败: {}",
        'embed_status_run': "嵌入: {}/{} (成功: {} | 失败: {})",
        # Shared
        'btn_start': "开始处理",
        'btn_retry_failed': "重试失败",
        'btn_stop': "停止",
        'status_idle': "准备就绪",
        'status_run': "进度: {}/{} (成功: {} | 失败: {} | 跳过: {})",
        'status_stop': "正在停止...",
        'log_ready': "已就绪。请加载含图片的表格文件。",
        'log_header': "✅ 锁定表头: 第 {} 行",
        'log_stats': "📊 列分析: 列 {} 含 {} 条有效数据 (类型: {})",
        'msg_404': "❌ {}: [404] 链接失效/不存在",
        'msg_timeout': "⚠️ {}: [超时] 网络连接卡顿",
        'msg_err': "❌ {}: [错误] {}",
        'menu_help': "帮助",
        'menu_check_update': "检查更新",
        'update_available': "⬆ 发现新版本 {}！点击下载",
        'update_none': "✅ 当前已是最新版本",
        'update_check_fail': "检查更新失败",
    },
    'en': {
        'title': "SheetPic - Spreadsheet Image Extract & Embed",
        'menu_lang': "Language",
        'footer_text': "Build by Andre  |  v{}",
        'tab_extract': "Extract",
        'tab_embed': "Embed",
        'sec_source': "Data Source",
        'sec_settings': "Settings",
        'sec_embed_settings': "Embed Settings",
        'btn_browse': "📂 File",
        'btn_clip': "📋 Clip",
        'lbl_dest': "Output:",
        'btn_dest': "Change",
        'lbl_sheet': "Sheet:",
        'unnamed': "Unnamed",
        # Extract
        'lbl_img': "Image Source (Auto Merge)",
        'lbl_code': "Filename Column",
        'opt_auto': "★ [Auto Merge] Priority by count",
        'type_url': "[Link] {} ({} URLs)",
        'type_img': "[Image] {} ({} Embedded)",
        'lbl_extract_bg': "Extract Background",
        'opt_extract_bg_original': "Original",
        'opt_extract_bg_white': "White",
        'lbl_extract_shape': "Background Shape",
        'opt_extract_shape_original': "Original Ratio",
        'opt_extract_shape_square': "Square",
        'msg_skip': "❌ {}: [Skip] No valid image found",
        'done_msg': "Time: {:.1f}s\nSuccess: {}\nFailed: {}\nSkipped: {}\nPath: {}",
        # Embed
        'lbl_url_col': "Image URL Column",
        'lbl_sku_col': "SKU/ID Column (for ordering)",
        'lbl_img_size': "Max Dimension (px)",
        'lbl_url_library': "URL Library: {} items",
        'lbl_url_library_fields': "URL Library Fields: {}",
        'btn_import_url_lib': "Import URL Library",
        'btn_clear_url_lib': "Clear URL Library",
        'btn_select_all_fields': "All",
        'btn_clear_fields': "None",
        'lbl_img_bg': "Image Background",
        'opt_bg_white': "White JPG",
        'opt_bg_transparent': "Preserve PNG alpha",
        'chk_original': "Original Size (no resize)",
        'chk_write_original': "Write to original file (preserve format)",
        'msg_no_url': "❌ No URL column detected",
        'opt_url_library': "[URL Library] Match by SKU/ID ({} items)",
        'msg_url_lib_imported': "✅ URL library imported: {} added/updated, {} total ({})",
        'msg_url_lib_empty': "❌ URL library file is empty or has no valid mappings",
        'msg_url_lib_no_cols': "❌ URL library needs one SKU/ID column and one image URL column",
        'msg_url_lib_cleared': "URL library cleared",
        'msg_url_lib_clear_confirm': "Clear the saved URL library?",
        'msg_use_url_library': "No URL column detected; using URL library by SKU/ID",
        'msg_url_lib_matches': "URL library matches: {} / {}",
        'msg_embed_done': "Time: {:.1f}s\nEmbedded: {}\nFailed: {}\nOutput: {}",
        'msg_dl_fail': "[Download Failed]",
        'msg_dl_skip': "[No URL]",
        'msg_same_name_skip': "⏭ {}: [Same Name Skip] File already exists",
        'msg_invalid_url': "⚠️ {}: [Invalid URL] {}",
        'msg_conn_err': "❌ {}: [Connection Error] {}",
        'msg_ssl_err': "❌ {}: [SSL Error] {}",
        'msg_too_large': "❌ {}: [File Too Large] {}MB",
        'msg_bad_image': "❌ {}: [Bad Image] {}",
        'log_embed_start': "Starting image embedding...",
        'log_embed_save': "Saving Excel file...",
        'log_embed_format_fallback': "{} cannot be written with preserved formatting; creating a new .xlsx instead.",
        'msg_embed_error': "❌ Embed failed: {}",
        'embed_status_run': "Embed: {} / {} (OK: {} | Fail: {})",
        # Shared
        'btn_start': "Start",
        'btn_retry_failed': "Retry Failed",
        'btn_stop': "Stop",
        'status_idle': "Ready",
        'status_run': "{} / {} (OK: {} Fail: {} Skip: {})",
        'status_stop': "Stopping...",
        'log_ready': "Ready. Load a table with images.",
        'log_header': "✅ Header at Row {}",
        'log_stats': "📊 Col Stats: {} has {} valid items ({})",
        'msg_404': "❌ {}: [404] Not Found",
        'msg_timeout': "⚠️ {}: [Timeout] Connection failed",
        'msg_err': "❌ {}: [Error] {}",
        'menu_help': "Help",
        'menu_check_update': "Check for Updates",
        'update_available': "⬆ New version {} available! Click to download",
        'update_none': "✅ Already up to date",
        'update_check_fail': "Update check failed",
    }
}

COLORS = {
    'bg': '#F0F0F0', 'card': '#FFFFFF', 'primary': '#2563EB', 'primary_hov': '#1D4ED8',
    'danger': '#DC2626', 'text': '#1F2937', 'text_sub': '#666666', 'success': '#10B981',
    'border': '#DDDDDD', 'disabled_bg': '#CCCCCC', 'disabled_fg': '#555555'
}

GITHUB_URL = "https://github.com/youngoris/SheetPic"
EMBED_IMAGE_BORDER_RATIO = 0.05
EMBED_IMAGE_JPEG_QUALITY = 85
EMBED_BG_WHITE = "white"
EMBED_BG_TRANSPARENT = "transparent"
EXTRACT_IMAGE_BORDER_RATIO = 0.05
EXTRACT_BG_ORIGINAL = "original"
EXTRACT_BG_WHITE = "white"
EXTRACT_SHAPE_ORIGINAL = "original"
EXTRACT_SHAPE_SQUARE = "square"
URL_LIBRARY_CONFIG_KEY = "url_library"
URL_LIBRARY_RECORDS_CONFIG_KEY = "url_library_records"
URL_LIBRARY_FIELDS_CONFIG_KEY = "url_library_fields"
URL_LIBRARY_SELECTED_FIELDS_CONFIG_KEY = "url_library_selected_fields"
EXTRACT_TIMEOUT_RETRIES = 2
URL_LIBRARY_CODE_KEYWORDS = (
    "条形码", "条码", "商品条码", "sku", "barcode", "bar code",
    "ean", "upc", "gtin", "货号", "商品编码", "编码", "编号",
    "code", "id",
)
SKU_COLUMN_EXCLUDE_KEYWORDS = ("货位", "库位", "货架", "层板", "位置")
SKU_COLUMN_HIGH_PRIORITY_KEYWORDS = (
    "条形码", "商品条码", "条码", "sku", "barcode", "bar code",
    "ean", "upc", "gtin",
)
SKU_COLUMN_MEDIUM_PRIORITY_KEYWORDS = (
    "货号", "商品编码", "产品编码", "商品编号", "产品编号",
    "code", "id", "编码", "编号",
)


# ==========================================
# Header-row detection (exposed for testing)
# ==========================================
HEADER_KEYWORDS = {
    # Chinese
    '图片', '图', '图像', '主图', '链接', '网址', '地址', '编号', '编码', '货号',
    '型号', '商品', '产品', '名称', '品名', '规格', '颜色', '尺寸', '尺码',
    '价格', '单价', '售价', '成本', '数量', '库存', '单位', '重量', '材质',
    '描述', '备注', '分类', '类目', '品牌', '日期', '时间', '订单', '客户',
    'sku', '条码', '条形码', '序号',
    # English
    'image', 'img', 'photo', 'picture', 'thumbnail', 'url', 'link', 'href',
    'id', 'code', 'sku', 'name', 'title', 'product', 'item', 'brand',
    'price', 'cost', 'qty', 'quantity', 'stock', 'size', 'color', 'colour',
    'weight', 'desc', 'description', 'note', 'remark', 'category', 'date',
    'time', 'order', 'customer', 'no', 'no.', 'number',
}


def _is_blank(v):
    if v is None:
        return True
    if isinstance(v, float):
        try:
            import math
            return math.isnan(v)
        except Exception:
            return False
    if isinstance(v, str):
        return v.strip() == ''
    return False


def _looks_like_url(s):
    if not isinstance(s, str):
        return False
    s = s.strip().lower()
    return s.startswith('http://') or s.startswith('https://') or s.startswith('//')


def _normalize_lookup_code(value):
    if _is_blank(value):
        return ''
    if isinstance(value, float):
        if math.isnan(value):
            return ''
        if value.is_integer():
            return str(int(value))
    text = str(value).strip()
    if text.lower() in ('nan', 'none'):
        return ''
    if re.fullmatch(r'\d+\.0+', text):
        return text.split('.')[0]
    return text


def _extract_http_url(value):
    if _is_blank(value):
        return None
    text = str(value).strip()
    if text.lower() in ('nan', 'none'):
        return None
    if text.startswith('//'):
        return 'https:' + text
    match = re.search(r'https?://[^\s;]+', text, flags=re.IGNORECASE)
    if match:
        return match.group(0)
    return None


def _normalize_library_field_name(value):
    text = str(value).strip()
    if not text or text.startswith("Unnamed"):
        return ''
    return text


def _json_safe_value(value):
    if _is_blank(value):
        return ''
    if hasattr(value, 'item'):
        try:
            value = value.item()
        except Exception:
            pass
    if isinstance(value, float) and math.isnan(value):
        return ''
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _cell_type(v):
    """Classify a cell into one of: blank/str/num/date/url."""
    if _is_blank(v):
        return 'blank'
    if isinstance(v, bool):
        return 'num'
    if isinstance(v, (int, float)):
        return 'num'
    try:
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return 'date'
    except Exception:
        pass
    if isinstance(v, str):
        if _looks_like_url(v):
            return 'url'
        # numeric-looking strings still count as numbers (e.g., "123")
        s = v.strip()
        try:
            float(s.replace(',', ''))
            return 'num'
        except Exception:
            pass
        return 'str'
    return 'str'


def _image_has_transparency(img):
    if img.mode in ('RGBA', 'LA'):
        return True
    return img.mode == 'P' and 'transparency' in img.info


def _prepare_embed_image_bytes(pil_img, max_dim=None, bg_mode=EMBED_BG_WHITE):
    """Prepare image bytes with a 5% border; preserve alpha only when source has it."""
    bg_mode = bg_mode if bg_mode in (EMBED_BG_WHITE, EMBED_BG_TRANSPARENT) else EMBED_BG_WHITE
    has_transparency = _image_has_transparency(pil_img)
    preserve_transparency = bg_mode == EMBED_BG_TRANSPARENT and has_transparency
    if max_dim and max_dim > 0:
        content_max_dim = max(1, int(max_dim / (1 + EMBED_IMAGE_BORDER_RATIO * 2)))
        pil_img.thumbnail((content_max_dim, content_max_dim), PILImage.LANCZOS)

    if preserve_transparency:
        content = pil_img.convert('RGBA')
    else:
        if has_transparency:
            rgba = pil_img.convert('RGBA')
            content = PILImage.new('RGB', rgba.size, 'white')
            content.paste(rgba, mask=rgba.getchannel('A'))
        else:
            content = pil_img.convert('RGB')

    border_w = max(1, math.ceil(content.width * EMBED_IMAGE_BORDER_RATIO))
    border_h = max(1, math.ceil(content.height * EMBED_IMAGE_BORDER_RATIO))
    if preserve_transparency:
        canvas = PILImage.new(
            'RGBA',
            (content.width + border_w * 2, content.height + border_h * 2),
            (255, 255, 255, 0)
        )
        canvas.alpha_composite(content, (border_w, border_h))
    else:
        canvas = PILImage.new(
            'RGB',
            (content.width + border_w * 2, content.height + border_h * 2),
            'white'
        )
        canvas.paste(content, (border_w, border_h))

    if max_dim and max_dim > 0 and max(canvas.size) > max_dim:
        canvas.thumbnail((max_dim, max_dim), PILImage.LANCZOS)

    buf = BytesIO()
    if preserve_transparency:
        canvas.save(buf, format='PNG', optimize=True)
    else:
        canvas.save(buf, format='JPEG', quality=EMBED_IMAGE_JPEG_QUALITY, optimize=True)
    buf.seek(0)
    return buf


def _prepare_extract_image_bytes(
    pil_img,
    bg_mode=EXTRACT_BG_ORIGINAL,
    shape=EXTRACT_SHAPE_ORIGINAL,
    add_border=None,
):
    """Prepare extracted images for optional white canvas, square shape, and border."""
    bg_mode = bg_mode if bg_mode in (EXTRACT_BG_ORIGINAL, EXTRACT_BG_WHITE) else EXTRACT_BG_ORIGINAL
    shape = shape if shape in (EXTRACT_SHAPE_ORIGINAL, EXTRACT_SHAPE_SQUARE) else EXTRACT_SHAPE_ORIGINAL
    if add_border is None:
        add_border = bg_mode == EXTRACT_BG_WHITE
    has_transparency = _image_has_transparency(pil_img)
    preserve_transparency = bg_mode == EXTRACT_BG_ORIGINAL and has_transparency

    if preserve_transparency:
        content = pil_img.convert('RGBA')
    else:
        if has_transparency:
            rgba = pil_img.convert('RGBA')
            content = PILImage.new('RGB', rgba.size, 'white')
            content.paste(rgba, mask=rgba.getchannel('A'))
        else:
            content = pil_img.convert('RGB')

    border_w = max(1, math.ceil(content.width * EXTRACT_IMAGE_BORDER_RATIO)) if add_border else 0
    border_h = max(1, math.ceil(content.height * EXTRACT_IMAGE_BORDER_RATIO)) if add_border else 0
    canvas_w = content.width + border_w * 2
    canvas_h = content.height + border_h * 2
    if shape == EXTRACT_SHAPE_SQUARE:
        side = max(canvas_w, canvas_h)
        canvas_w = side
        canvas_h = side

    if preserve_transparency:
        canvas = PILImage.new('RGBA', (canvas_w, canvas_h), (255, 255, 255, 0))
    else:
        canvas = PILImage.new('RGB', (canvas_w, canvas_h), 'white')

    x = (canvas_w - content.width) // 2
    y = (canvas_h - content.height) // 2
    if preserve_transparency and add_border:
        x0 = max(0, x - border_w)
        y0 = max(0, y - border_h)
        x1 = min(canvas_w, x + content.width + border_w)
        y1 = min(canvas_h, y + content.height + border_h)
        canvas.paste((255, 255, 255, 255), (x0, y0, x1, y))
        canvas.paste((255, 255, 255, 255), (x0, y + content.height, x1, y1))
        canvas.paste((255, 255, 255, 255), (x0, y, x, y + content.height))
        canvas.paste((255, 255, 255, 255), (x + content.width, y, x1, y + content.height))

    if preserve_transparency:
        canvas.alpha_composite(content, (x, y))
    else:
        canvas.paste(content, (x, y))

    buf = BytesIO()
    if preserve_transparency:
        canvas.save(buf, format='PNG', optimize=True)
        ext = ".png"
    else:
        canvas.save(buf, format='JPEG', quality=EMBED_IMAGE_JPEG_QUALITY, optimize=True)
        ext = ".jpg"
    buf.seek(0)
    return buf, ext


def _extract_options_require_processing(options):
    if not options:
        return False
    return (
        options.get('bg_mode') == EXTRACT_BG_WHITE
        or options.get('shape') == EXTRACT_SHAPE_SQUARE
        or bool(options.get('add_border'))
    )


def _row_signature(row):
    """Return (n_filled, type_counts dict, values list)."""
    vals = [v for v in row if not _is_blank(v)]
    types = {'str': 0, 'num': 0, 'date': 0, 'url': 0}
    for v in row:
        t = _cell_type(v)
        if t == 'blank':
            continue
        types[t] = types.get(t, 0) + 1
    return len(vals), types, vals


def _score_header_row(df_raw, scan_rows=15):
    """Score the first `scan_rows` rows of `df_raw` and return the best index.

    `df_raw` is a pandas DataFrame loaded with header=None.
    """
    import math as _math
    if df_raw is None or df_raw.empty:
        return 0
    n_total_rows = len(df_raw)
    n_cols = df_raw.shape[1]
    if n_cols == 0:
        return 0

    scan_rows = min(scan_rows, n_total_rows)
    # Widest non-trivial row width across the whole sample → expected col count
    row_widths = [_row_signature(df_raw.iloc[i].tolist())[0] for i in range(n_total_rows)]
    max_width = max(row_widths) if row_widths else 0
    if max_width == 0:
        return 0

    # Most common width across data-region rows (skip first 3 to avoid title bias)
    from collections import Counter
    tail_widths = [w for w in row_widths[3:] if w > 0]
    if tail_widths:
        mode_width = Counter(tail_widths).most_common(1)[0][0]
    else:
        mode_width = max_width

    expected_width = max(mode_width, int(max_width * 0.6))

    best_idx = 0
    best_score = -_math.inf

    for idx in range(scan_rows):
        row_vals = df_raw.iloc[idx].tolist()
        n_filled, types, vals = _row_signature(row_vals)
        if n_filled == 0:
            continue

        # Fill ratio relative to expected width
        fill_ratio = min(1.0, n_filled / expected_width) if expected_width else 0
        # String purity
        str_ratio = types['str'] / n_filled
        # No URLs in headers
        url_penalty = -0.5 if types['url'] > 0 else 0
        # Numbers in a "header" row are suspicious — but tolerated up to ~30%
        num_ratio = (types['num'] + types['date']) / n_filled

        # Uniqueness (case-insensitive, stripped)
        normalized = [str(v).strip().lower() for v in vals]
        uniq_ratio = len(set(normalized)) / len(normalized) if normalized else 0

        # Average label length — headers are short
        avg_len = sum(len(str(v)) for v in vals) / len(vals)
        # Penalize very long cells (likely descriptions/titles)
        len_score = 1.0 if avg_len <= 12 else max(0.0, 1.0 - (avg_len - 12) / 30.0)

        # Keyword match
        kw_hits = 0
        for v in vals:
            if not isinstance(v, str):
                continue
            low = v.strip().lower()
            if low in HEADER_KEYWORDS:
                kw_hits += 1
                continue
            # Substring match for common Chinese keywords
            for kw in HEADER_KEYWORDS:
                if len(kw) >= 2 and kw in low:
                    kw_hits += 1
                    break
        kw_score = min(1.0, kw_hits / max(1, n_filled))

        # "Followed by data" — look at next 3 rows: should be ≥ as wide and
        # contain MORE numbers/dates/URLs than this row (mixed types).
        followed_score = 0.0
        look = min(3, n_total_rows - idx - 1)
        if look > 0:
            wider_or_equal = 0
            more_mixed = 0
            for j in range(1, look + 1):
                nf, tt, _ = _row_signature(df_raw.iloc[idx + j].tolist())
                if nf >= max(1, n_filled - 1):
                    wider_or_equal += 1
                # data rows typically have more non-string content than the header
                if (tt['num'] + tt['date'] + tt['url']) > types['num'] + types['date'] + types['url']:
                    more_mixed += 1
            followed_score = (wider_or_equal / look) * 0.5 + (more_mixed / look) * 0.5
        else:
            # last row of the sheet can't be a header
            followed_score = -1.0

        # Sparse-row penalty (likely a merged title spanning few cells)
        sparse_penalty = -0.6 if n_filled < max(2, expected_width * 0.5) else 0.0

        score = (
            fill_ratio * 2.0 +
            str_ratio * 1.5 +
            uniq_ratio * 1.5 +
            len_score * 1.0 +
            kw_score * 1.5 +
            followed_score * 2.0 +
            url_penalty +
            sparse_penalty -
            num_ratio * 1.2
        )

        # Tie-breaker: prefer the LATER row (titles come first)
        if score > best_score + 1e-9 or (abs(score - best_score) <= 1e-9 and idx > best_idx):
            best_score = score
            best_idx = idx

    return best_idx


class SheetPicApp:
    def __init__(self, root):
        self.root = root
        self.setup_lang()
        self.root.title(f"{self.T['title']}  v{APP_VERSION}")
        self.root.configure(bg=COLORS['bg'])
        if platform.system() == "Darwin":
            self.root.geometry("620x750")
            self.root.minsize(620, 750)
        else:
            self.root.geometry("600x730")
            self.root.minsize(600, 730)

        self.default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        self.file_path = None
        self.df = None
        self.wb = None
        self.ws = None
        self.header_row = 0
        self.is_running = False
        self.sheet_names = []
        self._ui_thread = threading.current_thread()

        # Extract state
        self.sorted_img_cols = []
        self.var_extract_bg = None
        self.var_extract_shape = None

        # Embed state
        self.embed_url_col_idx = 0
        self.embed_sku_col_idx = 0
        self.embed_url_cols = []
        self.embed_use_url_library = False
        self.url_library = self._load_url_library()
        self.url_library_records = self._load_url_library_records()
        self.url_library_field_names = self._load_url_library_fields()
        self.url_library_selected_fields = self._load_url_library_selected_fields()
        self._url_library_combo_value = None
        self.var_img_bg = None
        self.extract_failed_tasks = []

        self.setup_style()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.after(2000, lambda: self.check_update(auto=True))

    CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".sheetpic_config")

    def setup_lang(self):
        # 1. 读取用户手动设置
        saved = self._load_config().get('lang')
        if saved in LANG_MAP:
            self.lang = saved
            self.T = LANG_MAP[self.lang]
            return

        # 2. 自动检测系统语言
        self.lang = self._detect_system_lang()
        self.T = LANG_MAP[self.lang]

    def _detect_system_lang(self):
        try:
            if platform.system() == "Darwin":
                import subprocess
                result = subprocess.run(
                    ['defaults', 'read', '-g', 'AppleLanguages'],
                    capture_output=True, text=True, timeout=3
                )
                if result.returncode == 0 and 'zh' in result.stdout.lower():
                    return 'zh'
            elif platform.system() == "Windows":
                # Windows API: GetUserDefaultUILanguage → LANGID
                import ctypes
                lang_id = ctypes.windll.kernel32.GetUserDefaultUILanguage()
                # 0x0804=zh-CN, 0x0404=zh-TW, 0x0C04=zh-HK, 0x1004=zh-SG
                if lang_id in (0x0804, 0x0404, 0x0C04, 0x1004):
                    return 'zh'
                # 也检查 MUI language list
                buf = ctypes.create_unicode_buffer(256)
                ctypes.windll.kernel32.GetUserPreferredUILanguages(
                    0x08, None, buf, ctypes.byref(ctypes.c_uint(256)))
                if 'zh' in buf.value.lower():
                    return 'zh'
            else:
                for var in ('LANG', 'LC_ALL', 'LC_MESSAGES'):
                    val = os.environ.get(var, '')
                    if 'zh' in val.lower():
                        return 'zh'
        except Exception:
            pass
        return 'en'

    def _load_config(self):
        try:
            with open(self.CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            return {}

    def _write_config(self, cfg):
        with open(self.CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False)

    def _save_config(self, lang):
        cfg = self._load_config()
        cfg['lang'] = lang
        self._write_config(cfg)

    def _load_url_library(self):
        raw = self._load_config().get(URL_LIBRARY_CONFIG_KEY, {})
        if not isinstance(raw, dict):
            return {}
        library = {}
        for code, url in raw.items():
            key = _normalize_lookup_code(code)
            clean_url = _extract_http_url(url)
            if key and clean_url:
                library[key] = clean_url
        return library

    def _load_url_library_records(self):
        raw = self._load_config().get(URL_LIBRARY_RECORDS_CONFIG_KEY, {})
        if not isinstance(raw, dict):
            return {}
        records = {}
        for code, record in raw.items():
            key = _normalize_lookup_code(code)
            if not key or not isinstance(record, dict):
                continue
            clean_record = {}
            for field, value in record.items():
                field_name = _normalize_library_field_name(field)
                if field_name:
                    clean_record[field_name] = _json_safe_value(value)
            if clean_record:
                records[key] = clean_record
        return records

    def _load_url_library_fields(self):
        raw = self._load_config().get(URL_LIBRARY_FIELDS_CONFIG_KEY, [])
        if not isinstance(raw, list):
            return []
        fields = []
        seen = set()
        for field in raw:
            field_name = _normalize_library_field_name(field)
            if field_name and field_name not in seen:
                fields.append(field_name)
                seen.add(field_name)
        return fields

    def _load_url_library_selected_fields(self):
        raw = self._load_config().get(URL_LIBRARY_SELECTED_FIELDS_CONFIG_KEY, [])
        if not isinstance(raw, list):
            return []
        available = set(getattr(self, 'url_library_field_names', []) or [])
        selected = []
        seen = set()
        for field in raw:
            field_name = _normalize_library_field_name(field)
            if field_name and field_name not in seen and (not available or field_name in available):
                selected.append(field_name)
                seen.add(field_name)
        return selected

    def _save_url_library(self):
        cfg = self._load_config()
        cfg[URL_LIBRARY_CONFIG_KEY] = getattr(self, 'url_library', {})
        cfg[URL_LIBRARY_RECORDS_CONFIG_KEY] = getattr(self, 'url_library_records', {})
        cfg[URL_LIBRARY_FIELDS_CONFIG_KEY] = getattr(self, 'url_library_field_names', [])
        cfg[URL_LIBRARY_SELECTED_FIELDS_CONFIG_KEY] = getattr(self, 'url_library_selected_fields', [])
        self._write_config(cfg)

    def switch_lang(self, lang):
        if lang == self.lang:
            return
        self.lang = lang
        self.T = LANG_MAP[lang]
        self._save_config(lang)
        messagebox.showinfo(
            self.T['title'],
            "语言已切换，程序将重启以应用更改。\nLanguage changed, restarting..." if lang == 'zh'
            else "Language changed. The app will restart to apply.\n语言已切换，程序将重启。"
        )
        self.root.destroy()
        os.execl(sys.executable, sys.executable, *sys.argv)

    def setup_style(self):
        style = ttk.Style()
        style.theme_use('clam')
        if platform.system() == "Darwin":
            base_font = ("PingFang SC", 11)
            bold_font = ("PingFang SC", 11, "bold")
        else:
            base_font = ("Microsoft YaHei UI", 11)
            bold_font = ("Microsoft YaHei UI", 11, "bold")

        style.configure(".", background=COLORS['card'], foreground=COLORS['text'], font=base_font)
        style.configure("TFrame", background=COLORS['card'])
        style.configure("TEntry", fieldbackground="#F5F5F5", bordercolor=COLORS['border'], padding=5)
        style.configure("TButton", background="#E8E8E8", foreground=COLORS['text'], borderwidth=0, font=base_font)
        style.map("TButton", background=[('active', '#D8D8D8'), ('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])
        style.configure("Primary.TButton", background=COLORS['primary'], foreground="white",
                         font=bold_font, borderwidth=0)
        style.map("Primary.TButton", background=[('active', COLORS['primary_hov']), ('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])
        style.configure("Danger.TButton", background=COLORS['danger'], foreground="white",
                         font=bold_font, borderwidth=0)
        style.map("Danger.TButton", background=[('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])
        style.configure("Green.Horizontal.TProgressbar", background=COLORS['success'],
                         troughcolor="#DDDDDD", bordercolor=COLORS['card'], thickness=6)
        # Notebook tab style
        style.configure("TNotebook", background="#FFFFFF", borderwidth=0)
        style.configure("TNotebook.Tab", font=base_font, padding=[16, 6],
                         background="#E0E0E0", foreground="#555555")
        style.map("TNotebook.Tab",
                   background=[("selected", "#FFFFFF"), ("!selected", "#E0E0E0")],
                   foreground=[("selected", "#1F2937"), ("!selected", "#888888")])

    def setup_ui(self):
        # === 菜单栏 ===
        menubar = tk.Menu(self.root)
        lang_menu = tk.Menu(menubar, tearoff=0)
        lang_menu.add_command(label="中文", command=lambda: self.switch_lang('zh'))
        lang_menu.add_command(label="English", command=lambda: self.switch_lang('en'))
        menubar.add_cascade(label=self.T['menu_lang'], menu=lang_menu)
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label=self.T['menu_check_update'], command=lambda: self.check_update(auto=False))
        menubar.add_cascade(label=self.T['menu_help'], menu=help_menu)
        self.root.config(menu=menubar)

        # === card1: 数据来源 (共享) ===
        card1 = tk.Frame(self.root, bg=COLORS['card'], padx=15, pady=15)
        card1.pack(fill='x', padx=15, pady=(20, 5))
        tk.Label(card1, text=self.T['sec_source'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))
        row1 = tk.Frame(card1, bg=COLORS['card'])
        row1.pack(fill='x')
        self.entry_path = ttk.Entry(row1)
        self.entry_path.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=3)
        self._setup_dnd(self.entry_path)
        ttk.Button(row1, text=self.T['btn_browse'], width=10, command=self.select_file).pack(side='left', padx=2)
        ttk.Button(row1, text=self.T['btn_clip'], width=8, command=self.load_clipboard).pack(side='left')

        # 工作表选择 (xlsx多Sheet)
        row_sheet = tk.Frame(card1, bg=COLORS['card'])
        row_sheet.pack(fill='x', pady=(10, 0))
        tk.Label(row_sheet, text=self.T['lbl_sheet'], bg=COLORS['card'], width=8, anchor='w').pack(side='left')
        self.combo_sheet = ttk.Combobox(row_sheet, state="disabled")
        self.combo_sheet.pack(side='left', fill='x', expand=True, padx=5, ipady=3)
        self.combo_sheet.bind('<<ComboboxSelected>>', self.on_sheet_changed)

        # 输出目录 (共享)
        row_dest = tk.Frame(card1, bg=COLORS['card'])
        row_dest.pack(fill='x', pady=(10, 0))
        tk.Label(row_dest, text=self.T['lbl_dest'], bg=COLORS['card'], width=8, anchor='w').pack(side='left')
        self.entry_dest = ttk.Entry(row_dest)
        self.entry_dest.insert(0, self.default_dir)
        self.entry_dest.pack(side='left', fill='x', expand=True, padx=5, ipady=3)
        ttk.Button(row_dest, text=self.T['btn_dest'], width=6, command=self.select_folder).pack(side='left')

        # === Notebook: 两个 Tab ===
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='x', padx=15, pady=10)

        tab_extract = tk.Frame(self.notebook, bg=COLORS['card'], padx=15, pady=15)
        tab_embed = tk.Frame(self.notebook, bg=COLORS['card'], padx=15, pady=15)
        self.notebook.add(tab_extract, text=f"  {self.T['tab_extract']}  ")
        self.notebook.add(tab_embed, text=f"  {self.T['tab_embed']}  ")

        self.notebook.bind('<<NotebookTabChanged>>', self.on_tab_changed)

        # --- Tab 1: 提取图片 ---
        self._build_extract_tab(tab_extract)

        # --- Tab 2: 嵌入图片 ---
        self._build_embed_tab(tab_embed)

        # === 动作区 (共享) ===
        action_frame = tk.Frame(self.root, bg=COLORS['bg'])
        action_frame.pack(fill='x', padx=15, pady=5)
        self.progress = ttk.Progressbar(action_frame, orient="horizontal", mode="determinate",
                                         style="Green.Horizontal.TProgressbar")
        self.progress.pack(fill='x', pady=(0, 5))
        self.lbl_status = tk.Label(action_frame, text="...", bg=COLORS['bg'],
                                    fg=COLORS['text_sub'], font=("Arial", 10))
        self.lbl_status.pack(anchor='e')
        btn_box = tk.Frame(action_frame, bg=COLORS['bg'])
        btn_box.pack(fill='x', pady=5)
        self.btn_run = ttk.Button(btn_box, text=self.T['btn_start'], style="Primary.TButton",
                                   command=self.start_thread, state='disabled')
        self.btn_run.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=5)
        self.btn_retry = ttk.Button(btn_box, text=self.T['btn_retry_failed'],
                                    command=self.retry_failed_extract, state='disabled')
        self.btn_retry.pack(side='left', fill='x', expand=True, padx=5, ipady=5)
        self.btn_stop = ttk.Button(btn_box, text=self.T['btn_stop'], style="Danger.TButton",
                                    command=self.stop_thread, state='disabled')
        self.btn_stop.pack(side='right', fill='x', expand=True, padx=(5, 0), ipady=5)

        # === 页脚 (先打包，确保不被日志区挤掉) ===
        footer = tk.Frame(self.root, bg=COLORS['bg'])
        footer.pack(side='bottom', fill='x', padx=15, pady=8)
        tk.Label(footer, text=self.T['footer_text'].format(APP_VERSION),
                 font=("Arial", 10), bg=COLORS['bg'], fg=COLORS['text_sub']).pack(side='left')
        self.lbl_update = tk.Label(footer, text="", font=("Arial", 10, "bold"),
                                   bg=COLORS['bg'], fg=COLORS['primary'], cursor="hand2")
        self.lbl_update.pack(side='left', padx=(10, 0))
        lbl_link = tk.Label(footer, text="GitHub", font=("Arial", 10),
                             bg=COLORS['bg'], fg=COLORS['primary'], cursor="hand2")
        lbl_link.pack(side='right')
        lbl_link.bind("<Button-1>", lambda e: webbrowser.open(GITHUB_URL))

        # === 日志区 ===
        log_frame = tk.Frame(self.root, bg=COLORS['card'], bd=1, relief="flat")
        log_frame.pack(fill='both', expand=True, padx=15, pady=(5, 5))
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, font=("Consolas", 10),
                                                   bd=0, highlightthickness=0)
        self.log_text.pack(fill='both', expand=True)
        self.log_text.configure(bg="#F5F5F5", fg="#444", padx=10, pady=10, state='normal')

        self.mode = 'extract'
        self.log(self.T['log_ready'])

    def _build_extract_tab(self, parent):
        """构建「提取图片」Tab"""
        tk.Label(parent, text=self.T['sec_settings'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))

        # 上下两行布局
        tk.Label(parent, text=self.T['lbl_img'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.combo_img = ttk.Combobox(parent, state="disabled")
        self.combo_img.pack(fill='x', pady=(2, 6))

        tk.Label(parent, text=self.T['lbl_code'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.combo_code = ttk.Combobox(parent, state="disabled")
        self.combo_code.pack(fill='x', pady=(2, 0))

        row_options = tk.Frame(parent, bg=COLORS['card'])
        row_options.pack(fill='x', pady=(10, 0))

        bg_frame = tk.Frame(row_options, bg=COLORS['card'])
        bg_frame.pack(side='left', padx=(0, 18), anchor='n')
        tk.Label(bg_frame, text=self.T['lbl_extract_bg'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.var_extract_bg = tk.StringVar(value=EXTRACT_BG_ORIGINAL)
        tk.Radiobutton(bg_frame, text=self.T['opt_extract_bg_original'], variable=self.var_extract_bg,
                       value=EXTRACT_BG_ORIGINAL, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(anchor='w')
        tk.Radiobutton(bg_frame, text=self.T['opt_extract_bg_white'], variable=self.var_extract_bg,
                       value=EXTRACT_BG_WHITE, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(anchor='w')

        shape_frame = tk.Frame(row_options, bg=COLORS['card'])
        shape_frame.pack(side='left', padx=(0, 18), anchor='n')
        tk.Label(shape_frame, text=self.T['lbl_extract_shape'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.var_extract_shape = tk.StringVar(value=EXTRACT_SHAPE_ORIGINAL)
        tk.Radiobutton(shape_frame, text=self.T['opt_extract_shape_original'], variable=self.var_extract_shape,
                       value=EXTRACT_SHAPE_ORIGINAL, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(anchor='w')
        tk.Radiobutton(shape_frame, text=self.T['opt_extract_shape_square'], variable=self.var_extract_shape,
                       value=EXTRACT_SHAPE_SQUARE, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(anchor='w')

    def _build_embed_tab(self, parent):
        """构建「嵌入图片」Tab"""
        tk.Label(parent, text=self.T['sec_embed_settings'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10, "bold")).pack(anchor='w', pady=(0, 5))

        # URL列
        row_url = tk.Frame(parent, bg=COLORS['card'])
        row_url.pack(fill='x', pady=(0, 8))
        tk.Label(row_url, text=self.T['lbl_url_col'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.combo_url = ttk.Combobox(row_url, state="disabled")
        self.combo_url.pack(fill='x', pady=(2, 0))

        # SKU列
        row_sku = tk.Frame(parent, bg=COLORS['card'])
        row_sku.pack(fill='x', pady=(0, 8))
        tk.Label(row_sku, text=self.T['lbl_sku_col'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.combo_sku = ttk.Combobox(row_sku, state="disabled")
        self.combo_sku.pack(fill='x', pady=(2, 0))

        # URL库
        row_lib = tk.Frame(parent, bg=COLORS['card'])
        row_lib.pack(fill='x', pady=(0, 8))
        self.lbl_url_library = tk.Label(
            row_lib,
            text=self.T['lbl_url_library'].format(len(getattr(self, 'url_library', {}))),
            bg=COLORS['card'],
            fg=COLORS['text_sub'],
            font=("Arial", 10)
        )
        self.lbl_url_library.pack(side='left')
        ttk.Button(row_lib, text=self.T['btn_import_url_lib'],
                   command=self.import_url_library).pack(side='right', padx=(5, 0))
        ttk.Button(row_lib, text=self.T['btn_clear_url_lib'],
                   command=self.clear_url_library).pack(side='right')

        row_fields = tk.Frame(parent, bg=COLORS['card'])
        row_fields.pack(fill='x', pady=(0, 8))
        fields_head = tk.Frame(row_fields, bg=COLORS['card'])
        fields_head.pack(fill='x')
        self.lbl_url_library_fields = tk.Label(
            fields_head,
            text=self.T['lbl_url_library_fields'].format(len(getattr(self, 'url_library_field_names', []))),
            bg=COLORS['card'],
            fg=COLORS['text_sub'],
            font=("Arial", 10)
        )
        self.lbl_url_library_fields.pack(side='left')
        self.btn_url_fields_all = ttk.Button(
            fields_head,
            text=self.T['btn_select_all_fields'],
            command=self.select_all_url_library_fields
        )
        self.btn_url_fields_all.pack(side='right', padx=(5, 0))
        self.btn_url_fields_none = ttk.Button(
            fields_head,
            text=self.T['btn_clear_fields'],
            command=self.clear_selected_url_library_fields
        )
        self.btn_url_fields_none.pack(side='right')

        fields_box = tk.Frame(row_fields, bg=COLORS['card'])
        fields_box.pack(fill='x', pady=(2, 0))
        self.list_url_fields = tk.Listbox(
            fields_box,
            height=4,
            selectmode=tk.MULTIPLE,
            exportselection=False,
            relief='solid',
            borderwidth=1,
            highlightthickness=0
        )
        self.list_url_fields.pack(side='left', fill='x', expand=True)
        fields_scroll = ttk.Scrollbar(fields_box, orient='vertical', command=self.list_url_fields.yview)
        fields_scroll.pack(side='right', fill='y')
        self.list_url_fields.config(yscrollcommand=fields_scroll.set)
        self.list_url_fields.bind('<<ListboxSelect>>', self._on_url_library_fields_selected)
        self._refresh_url_library_fields_ui()

        # 最大边长 + 选项 同行排列
        row_size = tk.Frame(parent, bg=COLORS['card'])
        row_size.pack(fill='x')
        tk.Label(row_size, text=self.T['lbl_img_size'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')

        size_frame = tk.Frame(parent, bg=COLORS['card'])
        size_frame.pack(fill='x', pady=(2, 0))
        self.entry_max_dim = ttk.Entry(size_frame, width=8)
        self.entry_max_dim.grid(row=0, column=0, rowspan=2, padx=(0, 14), sticky='nw', pady=(2, 0))
        self.entry_max_dim.insert(0, "500")

        def _mk_chk(text, var, **kw):
            return tk.Checkbutton(size_frame, text=text, variable=var,
                                  bg=COLORS['card'], fg=COLORS['text_sub'],
                                  font=("Arial", 10),
                                  activebackground=COLORS['card'], **kw)

        self.var_original = tk.BooleanVar(value=False)
        chk_original = _mk_chk(self.T['chk_original'], self.var_original,
                               command=self._toggle_max_dim)
        chk_original.grid(row=0, column=1, sticky='w', padx=(0, 12))

        self.var_write_original = tk.BooleanVar(value=False)
        chk_wo = _mk_chk(self.T['chk_write_original'], self.var_write_original)
        chk_wo.grid(row=0, column=2, sticky='w')

        bg_frame = tk.Frame(size_frame, bg=COLORS['card'])
        bg_frame.grid(row=0, column=3, rowspan=2, sticky='w', padx=(16, 0))
        tk.Label(bg_frame, text=self.T['lbl_img_bg'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 10)).pack(anchor='w')
        self.var_img_bg = tk.StringVar(value=EMBED_BG_WHITE)
        tk.Radiobutton(bg_frame, text=self.T['opt_bg_white'], variable=self.var_img_bg,
                       value=EMBED_BG_WHITE, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(side='left')
        tk.Radiobutton(bg_frame, text=self.T['opt_bg_transparent'], variable=self.var_img_bg,
                       value=EMBED_BG_TRANSPARENT, bg=COLORS['card'], fg=COLORS['text_sub'],
                       font=("Arial", 10), activebackground=COLORS['card']).pack(side='left')

    # ==========================================
    # 通用方法
    # ==========================================

    def _setup_dnd(self, widget):
        """Enable drag-and-drop on a widget if tkdnd is available."""
        try:
            widget.drop_target_register('DND_Files')
            widget.dnd_bind('<<Drop>>', self._on_drop)
        except (tk.TclError, AttributeError):
            pass

    def _on_drop(self, event):
        """Handle dropped files."""
        path = event.data.strip()
        # tkdnd may wrap paths in braces or quote them
        if path.startswith('{') and path.endswith('}'):
            path = path[1:-1]
        if path.startswith('"') and path.endswith('"'):
            path = path[1:-1]
        # Handle multiple files: take the first one
        if ' ' in path and not os.path.exists(path):
            # Try splitting by spaces, find first existing file
            for part in path.split():
                if os.path.exists(part):
                    path = part
                    break
        if os.path.exists(path):
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, path)
            self.file_path = path
            self.combo_sheet.set('')
            self.combo_sheet['values'] = []
            self.combo_sheet.config(state='disabled')
            threading.Thread(target=self.analyze_data, daemon=True).start()

    def on_tab_changed(self, event):
        tab = self.notebook.index(self.notebook.select())
        self.mode = 'extract' if tab == 0 else 'embed'
        self._update_retry_button_state()

    def _update_sheet_combo(self, selected):
        self.combo_sheet['values'] = self.sheet_names
        if selected in self.sheet_names:
            self.combo_sheet.set(selected)
        elif self.sheet_names:
            self.combo_sheet.current(0)
        if len(self.sheet_names) > 1:
            self.combo_sheet.config(state='readonly')
        else:
            self.combo_sheet.config(state='disabled')

    def on_sheet_changed(self, event=None):
        if not self.wb:
            return
        name = self.combo_sheet.get()
        if name not in self.wb.sheetnames:
            return
        self.ws = self.wb[name]
        self.log(f">>> Sheet: {name}")
        threading.Thread(target=self._reload_sheet_data, daemon=True).start()

    def _reload_sheet_data(self):
        self.root.after(0, lambda: self.progress.config(mode='indeterminate'))
        self.root.after(0, lambda: self.progress.start(15))
        self.df = None
        self.header_row = 0
        try:
            selected_sheet = self.combo_sheet.get()
            self.header_row = self.find_robust_header(self.file_path, sheet_name=selected_sheet)
            if self.header_row > 0:
                self.root.after(0, lambda: self.log(self.T['log_header'].format(self.header_row + 1)))
            self.df = pd.read_excel(self.file_path, header=self.header_row, sheet_name=selected_sheet)
        except Exception as e:
            self.root.after(0, lambda: self.log(f"❌ Error: {e}"))
        self.root.after(0, lambda: self.progress.stop())
        self.root.after(0, lambda: self.progress.config(mode='determinate'))
        self.root.after(0, lambda: self.progress.__setitem__('value', 0))
        if self.df is not None and not self.df.empty:
            self.process_df()

    def log(self, msg):
        if getattr(self, '_ui_thread', None) is not threading.current_thread():
            try:
                self.root.after(0, self.log, msg)
            except (RuntimeError, tk.TclError):
                pass
            return

        now = datetime.datetime.now().strftime("[%H:%M:%S]")
        previous_state = None
        try:
            previous_state = self.log_text.cget('state')
        except (tk.TclError, AttributeError):
            pass
        try:
            if previous_state == 'disabled':
                self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, f"{now} {msg}\n")
            self.log_text.see(tk.END)
        finally:
            if previous_state == 'disabled':
                try:
                    self.log_text.configure(state='disabled')
                except (tk.TclError, AttributeError):
                    pass

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx;*.xls;*.csv;*.html"), ("All", "*.*")])
        if p:
            self.file_path = p
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, os.path.basename(p))
            self.combo_sheet.set('')
            self.combo_sheet['values'] = []
            self.combo_sheet.config(state='disabled')
            threading.Thread(target=self.analyze_data, daemon=True).start()

    def select_folder(self):
        d = filedialog.askdirectory()
        if d:
            self.entry_dest.delete(0, tk.END)
            self.entry_dest.insert(0, d)

    def import_url_library(self):
        path = filedialog.askopenfilename(
            filetypes=[("Data", "*.xlsx;*.xls;*.csv;*.html"), ("All", "*.*")]
        )
        if not path:
            return
        try:
            df = self._read_table_for_url_library(path)
            added = self._merge_url_library_from_df(df)
            if added is None:
                self.log(self.T['msg_url_lib_no_cols'])
                return
            if added == 0:
                self.log(self.T['msg_url_lib_empty'])
                return
            self._save_url_library()
            self._refresh_url_library_status()
            self.log(self.T['msg_url_lib_imported'].format(
                added, len(self.url_library), os.path.basename(path)
            ))
            if self.df is not None and not self.df.empty:
                self.process_df()
        except Exception as e:
            self.log(f"❌ Error: {e}")

    def clear_url_library(self):
        if not getattr(self, 'url_library', {}):
            return
        if not messagebox.askyesno(self.T['title'], self.T['msg_url_lib_clear_confirm']):
            return
        self.url_library = {}
        self.url_library_records = {}
        self.url_library_field_names = []
        self.url_library_selected_fields = []
        self._save_url_library()
        self._refresh_url_library_status()
        self.log(self.T['msg_url_lib_cleared'])
        if self.df is not None and not self.df.empty:
            self.process_df()

    def _refresh_url_library_status(self):
        if hasattr(self, 'lbl_url_library'):
            self.lbl_url_library.config(
                text=self.T['lbl_url_library'].format(len(getattr(self, 'url_library', {})))
            )
        self._refresh_url_library_fields_ui()

    def _refresh_url_library_fields_ui(self):
        if not hasattr(self, 'list_url_fields'):
            return

        fields = list(getattr(self, 'url_library_field_names', []) or [])
        selected = set(getattr(self, 'url_library_selected_fields', []) or [])
        self._updating_url_field_selection = True
        try:
            self.list_url_fields.config(state='normal')
            self.list_url_fields.delete(0, tk.END)
            for field in fields:
                self.list_url_fields.insert(tk.END, field)
            for idx, field in enumerate(fields):
                if field in selected:
                    self.list_url_fields.selection_set(idx)
        finally:
            self._updating_url_field_selection = False

        state = 'normal' if fields else 'disabled'
        self.list_url_fields.config(state=state)
        if hasattr(self, 'btn_url_fields_all'):
            self.btn_url_fields_all.config(state=state)
        if hasattr(self, 'btn_url_fields_none'):
            self.btn_url_fields_none.config(state=state)
        if hasattr(self, 'lbl_url_library_fields'):
            self.lbl_url_library_fields.config(text=self.T['lbl_url_library_fields'].format(len(fields)))

    def _on_url_library_fields_selected(self, _event=None):
        if getattr(self, '_updating_url_field_selection', False):
            return
        self.url_library_selected_fields = self._get_selected_url_library_fields(read_widget=True)
        self._save_url_library()

    def _get_selected_url_library_fields(self, read_widget=False):
        fields = list(getattr(self, 'url_library_field_names', []) or [])
        if read_widget and hasattr(self, 'list_url_fields'):
            selected_indices = set(int(i) for i in self.list_url_fields.curselection())
            selected = [field for idx, field in enumerate(fields) if idx in selected_indices]
        else:
            selected_set = set(getattr(self, 'url_library_selected_fields', []) or [])
            selected = [field for field in fields if field in selected_set]
        return selected

    def select_all_url_library_fields(self):
        self.url_library_selected_fields = list(getattr(self, 'url_library_field_names', []) or [])
        self._save_url_library()
        self._refresh_url_library_fields_ui()

    def clear_selected_url_library_fields(self):
        self.url_library_selected_fields = []
        self._save_url_library()
        self._refresh_url_library_fields_ui()

    def _read_table_for_url_library(self, path):
        ext = os.path.splitext(path)[1].lower()
        if ext == '.csv':
            try:
                return pd.read_csv(path, encoding='utf-8-sig', on_bad_lines='skip')
            except Exception:
                return pd.read_csv(path, encoding='gbk', on_bad_lines='skip')
        if ext == '.html':
            return pd.read_html(path)[0]
        header_row = self.find_robust_header(path)
        return pd.read_excel(path, header=header_row)

    def _detect_url_library_columns(self, df):
        code_col_indices, url_col_idx = self._detect_url_library_mapping_columns(df)
        first_code_idx = code_col_indices[0] if code_col_indices else None
        return first_code_idx, url_col_idx

    @staticmethod
    def _is_code_like_column_name(col_name):
        name = str(col_name).strip().lower()
        if not name or name.startswith("unnamed"):
            return False
        compact = re.sub(r"\s+", "", name)
        for keyword in URL_LIBRARY_CODE_KEYWORDS:
            key = keyword.lower().replace(" ", "")
            if key not in compact:
                continue
            if key == "id" and compact not in ("id", "商品id", "产品id") and not compact.endswith("id"):
                continue
            return True
        return False

    @staticmethod
    def _combo_option_column_name(option):
        return re.sub(r"\s+\([A-Z]+\)$", "", str(option)).strip()

    @staticmethod
    def _score_sku_column_name(col_name):
        name = str(col_name).strip().lower()
        if not name or name.startswith("unnamed"):
            return -1
        compact = re.sub(r"\s+", "", name)
        if any(keyword in compact for keyword in SKU_COLUMN_EXCLUDE_KEYWORDS):
            return -1
        for keyword in SKU_COLUMN_HIGH_PRIORITY_KEYWORDS:
            key = keyword.lower().replace(" ", "")
            if key in compact:
                return 200
        for keyword in SKU_COLUMN_MEDIUM_PRIORITY_KEYWORDS:
            key = keyword.lower().replace(" ", "")
            if key not in compact:
                continue
            if key == "id" and compact not in ("id", "商品id", "产品id") and not compact.endswith("id"):
                continue
            return 100
        return -1

    def _best_url_library_match_col_idx(self):
        library = getattr(self, 'url_library', {}) or {}
        if self.df is None or self.df.empty or not library:
            return None

        best_idx = None
        best_hits = 0
        best_name_score = -1
        for i, col_name in enumerate(self.df.columns):
            hits = 0
            for value in self.df.iloc[:, i]:
                code = _normalize_lookup_code(value)
                if code and code in library:
                    hits += 1
            name_score = self._score_sku_column_name(col_name)
            if hits > best_hits or (hits == best_hits and hits > 0 and name_score > best_name_score):
                best_idx = i
                best_hits = hits
                best_name_score = name_score

        return best_idx if best_hits > 0 else None

    def _detect_url_library_mapping_columns(self, df):
        if df is None or df.empty:
            return [], None

        url_counts = {}
        for i in range(len(df.columns)):
            count = self._count_http_values(df.iloc[:, i])
            if count > 0:
                url_counts[i] = count
        if not url_counts:
            return [], None
        url_col_idx = max(url_counts, key=url_counts.get)

        code_col_indices = []
        fallback_code_idx = None
        best_score = -1
        for i, col_name in enumerate(df.columns):
            if i == url_col_idx:
                continue
            series = df.iloc[:, i]
            nonblank_count = int(series.dropna().astype(str).str.strip().ne('').sum())
            if nonblank_count == 0:
                continue
            score = nonblank_count
            if self._is_code_like_column_name(col_name):
                score += 100000
                code_col_indices.append(i)
            if self._count_http_values(series) > 0:
                score -= 100000
            if score > best_score:
                best_score = score
                fallback_code_idx = i

        if not code_col_indices and fallback_code_idx is not None:
            code_col_indices = [fallback_code_idx]

        return code_col_indices, url_col_idx

    def _merge_url_library_from_df(self, df):
        code_col_indices, url_col_idx = self._detect_url_library_mapping_columns(df)
        if not code_col_indices or url_col_idx is None:
            return None

        if not hasattr(self, 'url_library') or self.url_library is None:
            self.url_library = {}
        if not hasattr(self, 'url_library_records') or self.url_library_records is None:
            self.url_library_records = {}
        if not hasattr(self, 'url_library_field_names') or self.url_library_field_names is None:
            self.url_library_field_names = []
        if not hasattr(self, 'url_library_selected_fields') or self.url_library_selected_fields is None:
            self.url_library_selected_fields = []

        field_columns = []
        existing_fields = list(self.url_library_field_names)
        existing_field_set = set(existing_fields)
        default_selected = set(self.url_library_selected_fields)
        code_col_set = set(code_col_indices)
        for idx, col_name in enumerate(df.columns):
            if idx == url_col_idx:
                continue
            field_name = _normalize_library_field_name(col_name)
            if not field_name:
                continue
            field_columns.append((idx, field_name))
            if field_name not in existing_field_set:
                existing_fields.append(field_name)
                existing_field_set.add(field_name)
                if idx not in code_col_set:
                    default_selected.add(field_name)
        self.url_library_field_names = existing_fields
        self.url_library_selected_fields = [
            field for field in self.url_library_field_names if field in default_selected
        ]

        added = 0
        for _, row in df.iterrows():
            url = _extract_http_url(row.iloc[url_col_idx])
            if not url:
                continue
            record = {}
            for field_col_idx, field_name in field_columns:
                record[field_name] = _json_safe_value(row.iloc[field_col_idx])
            row_codes = set()
            for code_col_idx in code_col_indices:
                code = _normalize_lookup_code(row.iloc[code_col_idx])
                if not code or code in row_codes:
                    continue
                row_codes.add(code)
                self.url_library[code] = url
                self.url_library_records[code] = record
                added += 1
        return added

    def load_clipboard(self):
        self.log(">>> Reading clipboard...")
        try:
            self.df = pd.read_clipboard()
            if not self.df.empty:
                self.file_path = "Clipboard"
                self.wb = None
                self.combo_sheet.set('')
                self.combo_sheet['values'] = []
                self.combo_sheet.config(state='disabled')
                self.entry_path.delete(0, tk.END)
                self.entry_path.insert(0, "Clipboard Data")
                self.process_df()
            else:
                self.log("❌ Clipboard empty")
        except Exception as e:
            self.log(f"❌ Error: {e}")

    def find_robust_header(self, file_path, sheet_name=0):
        """Locate the header row in an Excel sheet.

        Strategy: score the first ~15 rows on multiple signals and pick the
        best. A real header row is characterized by:
          - High fill ratio (close to the widest row in the sheet)
          - Mostly string cells with short labels
          - Unique non-null values (no duplicate column names)
          - Followed by data rows of comparable width but with mixed types
            (numbers / dates / mixed strings)
          - Bonus when cells contain common header keywords
          - Penalty when the row is sparse (likely a merged title)
        """
        try:
            if os.path.splitext(file_path)[1].lower() == '.csv':
                return 0
            df_raw = pd.read_excel(file_path, header=None, nrows=40, sheet_name=sheet_name)
            return _score_header_row(df_raw)
        except Exception:
            return 0

    def analyze_data(self):
        self.root.after(0, lambda: self.progress.config(mode='indeterminate'))
        self.root.after(0, lambda: self.progress.start(15))
        self.df = None
        if self.wb:
            try:
                self.wb.close()
            except Exception:
                pass
        self.wb = None
        self.ws = None
        self.header_row = 0

        try:
            ext = os.path.splitext(self.file_path)[1].lower() if self.file_path != "Clipboard" else ""

            selected_sheet = 0
            if ext == '.xlsx':
                try:
                    self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
                    self.ws = self.wb.active
                    self.sheet_names = self.wb.sheetnames
                    selected_sheet = self.ws.title
                    self.root.after(0, lambda: self._update_sheet_combo(selected_sheet))
                except Exception:
                    pass

            if ext in ['.xlsx', '.xls']:
                self.header_row = self.find_robust_header(self.file_path, sheet_name=selected_sheet)
                if self.header_row > 0:
                    self.log(self.T['log_header'].format(self.header_row + 1))

            if ext == '.csv':
                try:
                    self.df = pd.read_csv(self.file_path, encoding='utf-8-sig', on_bad_lines='skip')
                except Exception:
                    self.df = pd.read_csv(self.file_path, encoding='gbk', on_bad_lines='skip')
            elif ext == '.html':
                self.df = pd.read_html(self.file_path)[0]
            else:
                self.df = pd.read_excel(self.file_path, header=self.header_row, sheet_name=selected_sheet)

        except Exception as e:
            self.log(f"❌ Error: {e}")

        self.root.after(0, lambda: self.progress.stop())
        self.root.after(0, lambda: self.progress.config(mode='determinate'))
        self.root.after(0, lambda: self.progress.__setitem__('value', 0))
        if self.df is not None and not self.df.empty:
            self.process_df()

    def process_df(self):
        unnamed = self.T['unnamed']
        # Rename "Unnamed: N" placeholders, then de-duplicate so every column
        # has a unique label. Without this, duplicate names (common in real
        # spreadsheets, e.g. two "条码" columns or several blank header cells)
        # cause `df[name]` to return a DataFrame instead of a Series, breaking
        # `.str.contains` and silently hiding image/URL columns.
        new_cols = []
        seen = {}
        for c in self.df.columns:
            base = unnamed if str(c).startswith("Unnamed") else str(c)
            n = seen.get(base, 0)
            seen[base] = n + 1
            new_cols.append(base if n == 0 else f"{base}.{n}")
        self.df.columns = new_cols
        cols = list(self.df.columns)

        # --- Extract: 扫描嵌入图 + URL ---
        embed_counts = {}
        if self.wb:
            for img in getattr(self.ws, '_images', []):
                try:
                    c = img.anchor._from.col
                    embed_counts[c] = embed_counts.get(c, 0) + 1
                except (AttributeError, IndexError):
                    pass

        url_counts = {}
        for i in range(len(cols)):
            # Access by position to avoid duplicate-name pitfalls.
            series = self.df.iloc[:, i]
            real_count = self._count_http_values(series)
            if real_count > 0:
                url_counts[i] = real_count

        all_img_indices = set(embed_counts.keys()) | set(url_counts.keys())
        self.sorted_img_cols = []
        for idx in all_img_indices:
            count = max(embed_counts.get(idx, 0), url_counts.get(idx, 0))
            type_str = "embed" if idx in embed_counts else "url"
            self.sorted_img_cols.append({'idx': idx, 'count': count, 'type': type_str})
        self.sorted_img_cols.sort(key=lambda x: x['count'], reverse=True)

        # Extract combo 选项
        img_opts = []
        if self.sorted_img_cols:
            img_opts.append(self.T['opt_auto'])
        for item in self.sorted_img_cols:
            i = item['idx']
            col_letter = get_column_letter(i + 1)
            display_name = f"{cols[i]} ({col_letter})"
            if item['type'] == 'embed':
                label = self.T['type_img'].format(display_name, item['count'])
                self.log(self.T['log_stats'].format(col_letter, item['count'], "Embedded"))
            else:
                label = self.T['type_url'].format(display_name, item['count'])
                self.log(self.T['log_stats'].format(col_letter, item['count'], "URL"))
            img_opts.append(label)

        code_opts = [f"{c} ({get_column_letter(i+1)})" for i, c in enumerate(cols)]

        # --- Embed: URL列 + SKU列 ---
        self.embed_url_cols = [{'idx': i, 'count': c} for i, c in url_counts.items()]
        self.embed_url_cols.sort(key=lambda x: x['count'], reverse=True)

        url_opts = []
        for item in self.embed_url_cols:
            i = item['idx']
            col_letter = get_column_letter(i + 1)
            display_name = f"{cols[i]} ({col_letter})"
            url_opts.append(f"{display_name} - {item['count']} URLs")

        sku_opts = code_opts[:]  # same list

        self.root.after(0, lambda: self.update_ui_lists(img_opts, code_opts, url_opts, sku_opts))

    @staticmethod
    def _count_http_values(series):
        sample = series.dropna().head(50)
        if sample.empty:
            return 0
        sample_text = sample.astype(str)
        if not sample_text.str.contains("http", case=False, na=False, regex=False).any():
            return 0
        full_text = series.dropna().astype(str)
        return int(full_text.str.contains("http", case=False, na=False, regex=False).sum())

    def update_ui_lists(self, img_opts, code_opts, url_opts, sku_opts):
        # Extract combos
        self.combo_img['values'] = img_opts
        if img_opts:
            self.combo_img.current(0)
        self.combo_code['values'] = code_opts
        best = next((x for x in code_opts if any(k in x.lower() for k in ["code", "sku", "条码", "货号"])), None)
        if best:
            self.combo_code.set(best)
        elif code_opts:
            self.combo_code.current(0)
        self.combo_img.config(state='readonly')
        self.combo_code.config(state='readonly')

        # Embed combos
        library_count = len(getattr(self, 'url_library', {}))
        self._url_library_combo_value = None
        if library_count:
            self._url_library_combo_value = self.T['opt_url_library'].format(library_count)
            url_opts.append(self._url_library_combo_value)

        self.combo_url['values'] = url_opts
        self.embed_use_url_library = False
        if url_opts:
            self.combo_url.current(0)
            if self.embed_url_cols:
                self.embed_url_col_idx = self.embed_url_cols[0]['idx']
            else:
                self.embed_url_col_idx = None
                self.embed_use_url_library = True
                if self.mode == 'embed':
                    self.log(self.T['msg_use_url_library'])
        else:
            if self.mode == 'embed':
                self.log(self.T['msg_no_url'])

        self.combo_sku['values'] = sku_opts
        best_sku = None
        best_library_match_idx = self._best_url_library_match_col_idx() if library_count else None
        if best_library_match_idx is not None and best_library_match_idx < len(sku_opts):
            best_sku = sku_opts[best_library_match_idx]
        else:
            best_score = -1
            for option in sku_opts:
                score = self._score_sku_column_name(self._combo_option_column_name(option))
                if score > best_score:
                    best_sku = option
                    best_score = score
            if best_score < 0:
                best_sku = None

        if best_sku:
            self.combo_sku.set(best_sku)
            self.embed_sku_col_idx = self._get_col_index(best_sku)
        elif sku_opts:
            self.combo_sku.current(0)
            self.embed_sku_col_idx = 0
        self.combo_url.config(state='readonly' if url_opts else 'disabled')
        self.combo_sku.config(state='readonly')
        self.combo_url.bind('<<ComboboxSelected>>', self._on_url_selected)
        self.combo_sku.bind('<<ComboboxSelected>>', self._on_sku_selected)

        # Enable start button
        if img_opts or url_opts:
            self.btn_run.config(state='normal')
        else:
            self.btn_run.config(state='disabled')

    def _get_col_index(self, s):
        match = re.search(r'\(([A-Z]+)\)', s)
        if match:
            return column_index_from_string(match.group(1)) - 1
        return 0

    def _on_url_selected(self, event):
        value = self.combo_url.get()
        self.embed_use_url_library = bool(
            getattr(self, '_url_library_combo_value', None)
            and value == self._url_library_combo_value
        )
        self.embed_url_col_idx = None if self.embed_use_url_library else self._get_col_index(value)

    def _on_sku_selected(self, event):
        self.embed_sku_col_idx = self._get_col_index(self.combo_sku.get())

    def _toggle_max_dim(self):
        if self.var_original.get():
            self.entry_max_dim.config(state='disabled')
        else:
            self.entry_max_dim.config(state='normal')

    def _get_embed_bg_mode(self):
        var = getattr(self, 'var_img_bg', None)
        if var is None:
            return EMBED_BG_WHITE
        mode = var.get()
        if mode in (EMBED_BG_WHITE, EMBED_BG_TRANSPARENT):
            return mode
        return EMBED_BG_WHITE

    def _get_extract_image_options(self):
        bg_var = getattr(self, 'var_extract_bg', None)
        shape_var = getattr(self, 'var_extract_shape', None)
        bg_mode = bg_var.get() if bg_var is not None else EXTRACT_BG_ORIGINAL
        shape = shape_var.get() if shape_var is not None else EXTRACT_SHAPE_ORIGINAL
        add_border = bg_mode == EXTRACT_BG_WHITE
        options = {
            'bg_mode': bg_mode if bg_mode in (EXTRACT_BG_ORIGINAL, EXTRACT_BG_WHITE) else EXTRACT_BG_ORIGINAL,
            'shape': shape if shape in (EXTRACT_SHAPE_ORIGINAL, EXTRACT_SHAPE_SQUARE) else EXTRACT_SHAPE_ORIGINAL,
            'add_border': add_border,
        }
        return options if _extract_options_require_processing(options) else None

    # ==========================================
    # 线程控制
    # ==========================================

    def start_thread(self):
        self.is_running = True
        self.btn_run.config(state='disabled')
        if hasattr(self, 'btn_retry'):
            self.btn_retry.config(state='disabled')
        self.btn_stop.config(state='normal')
        if self.mode == 'extract':
            threading.Thread(target=self.run_extract_process, daemon=True).start()
        else:
            threading.Thread(target=self.run_embed_process, daemon=True).start()

    def _update_retry_button_state(self):
        if not hasattr(self, 'btn_retry'):
            return
        state = 'normal' if getattr(self, 'mode', 'extract') == 'extract' and getattr(self, 'extract_failed_tasks', []) else 'disabled'
        self.btn_retry.config(state=state)

    def retry_failed_extract(self):
        tasks = list(getattr(self, 'extract_failed_tasks', []) or [])
        if not tasks or self.is_running:
            return
        self.extract_failed_tasks = []
        self.is_running = True
        self._process_start_time = time.time()
        self.btn_run.config(state='disabled')
        self.btn_retry.config(state='disabled')
        self.btn_stop.config(state='normal')
        threading.Thread(target=self.run_extract_retry_process, args=(tasks,), daemon=True).start()

    def stop_thread(self):
        self.is_running = False
        self.log(">>> Stopping...")
        self.btn_stop.config(state='disabled')
        self.lbl_status.config(text=self.T['status_stop'])
        self.progress.stop()

    # ==========================================
    # 提取图片处理
    # ==========================================

    def run_extract_process(self):
        t_start = time.time()
        self._process_start_time = t_start
        self.extract_failed_tasks = []
        self.root.after(0, self._update_retry_button_state)
        dest = self.entry_dest.get()
        fname = "Clipboard" if self.file_path == "Clipboard" else os.path.splitext(os.path.basename(self.file_path))[0]
        out_dir = os.path.join(dest, f"{fname}_Img")
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)

        idx_code = self._get_col_index(self.combo_code.get())
        selection = self.combo_img.get()
        target_cols = []

        if "★" in selection:
            target_cols = self.sorted_img_cols
        else:
            sel_idx = self._get_col_index(selection)
            for item in self.sorted_img_cols:
                if item['idx'] == sel_idx:
                    target_cols = [item]
                    break

        img_map_row_col = {}
        if self.wb:
            for img in getattr(self.ws, '_images', []):
                r = img.anchor._from.row
                c = img.anchor._from.col
                if r not in img_map_row_col:
                    img_map_row_col[r] = {}
                img_map_row_col[r][c] = img

        success = 0
        fail = 0
        skipped = 0
        self.progress['maximum'] = len(self.df)
        tasks = {}
        planned_names = set()
        extract_options = self._get_extract_image_options()

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            for i in range(len(self.df)):
                if not self.is_running:
                    break

                code = str(self.df.iloc[i, idx_code]).strip()
                base_name = "".join([c for c in code if c.isalnum() or c in '-_'])
                if not base_name:
                    base_name = f"Row_{i+1}"

                row_images = []
                for col_info in target_cols:
                    c_idx = col_info['idx']
                    if col_info['type'] == 'embed':
                        excel_row = self.header_row + 1 + i
                        if excel_row in img_map_row_col and c_idx in img_map_row_col[excel_row]:
                            img_obj = img_map_row_col[excel_row][c_idx]
                            row_images.append(('embed', img_obj))
                    elif col_info['type'] == 'url':
                        val = str(self.df.iloc[i, c_idx]).strip()
                        if not val or val.lower() == 'nan' or "http" not in val.lower():
                            if val and val.lower() != 'nan':
                                self.root.after(0, lambda v=val: self.log(
                                    self.T['msg_invalid_url'].format(base_name, v[:60])))
                            continue
                        if not val.startswith("http"):
                            m = re.search(r'(https?://[^\s;]+)', val)
                            if m:
                                val = m.group(1)
                            else:
                                self.root.after(0, lambda v=val: self.log(
                                    self.T['msg_invalid_url'].format(base_name, v[:60])))
                                continue
                        val = self.clean_url(val.split('?')[0].split('!')[0])
                        row_images.append(('url', val))

                if not row_images:
                    skipped += 1
                    self.root.after(0, self.update_progress_ext, i+1+len(tasks), len(self.df), success, fail, skipped, self.T['msg_skip'].format(base_name))
                    continue

                for img_idx, (src_type, src_data) in enumerate(row_images):
                    suffix = f"-{img_idx}" if img_idx > 0 else ""
                    final_name = f"{base_name}{suffix}"
                    if final_name in planned_names or self._extract_output_exists(out_dir, final_name):
                        skipped += 1
                        self.root.after(0, self.update_progress_ext, i+1, len(self.df), success, fail, skipped,
                                        self.T['msg_same_name_skip'].format(final_name))
                        continue
                    planned_names.add(final_name)

                    if src_type == 'embed':
                        try:
                            raw_data = src_data._data()
                            if _extract_options_require_processing(extract_options):
                                is_ok, msg = self._save_processed_extract_image(
                                    raw_data,
                                    final_name,
                                    out_dir,
                                    extract_options
                                )
                                if is_ok:
                                    success += 1
                                else:
                                    fail += 1
                                    self.root.after(0, lambda m=msg: self.log(m))
                            else:
                                ext = ".png" if src_data.format == "png" else ".jpg"
                                path = os.path.join(out_dir, final_name + ext)
                                with open(path, "wb") as f:
                                    f.write(raw_data)
                                success += 1
                        except Exception as e:
                            fail += 1
                            self.root.after(0, lambda n=final_name, err=e: self.log(
                                self.T['msg_err'].format(n, f"{type(err).__name__}: {str(err)[:60]}")
                            ))
                    else:
                        task = {'url': src_data, 'filename_base': final_name, 'out_dir': out_dir}
                        if extract_options:
                            task['extract_options'] = extract_options
                            future = executor.submit(
                                self.download_url,
                                task['url'],
                                task['filename_base'],
                                task['out_dir'],
                                task['extract_options']
                            )
                        else:
                            future = executor.submit(
                                self.download_url,
                                task['url'],
                                task['filename_base'],
                                task['out_dir']
                            )
                        tasks[future] = task

                self.root.after(0, self.update_progress_ext, i+1, len(self.df), success, fail, skipped, "Process")

            done_count = i if 'i' in locals() else 0
            for future in concurrent.futures.as_completed(tasks):
                if not self.is_running:
                    break
                task = tasks[future]
                try:
                    is_ok, msg = future.result()
                except Exception as e:
                    is_ok, msg = False, self.T['msg_err'].format(
                        task['filename_base'],
                        f"{type(e).__name__}: {str(e)[:60]}"
                    )
                if is_ok:
                    success += 1
                else:
                    fail += 1
                    self.extract_failed_tasks.append(task)
                self.root.after(0, self.update_progress_ext, len(self.df), len(self.df), success, fail, skipped, msg)

        duration = time.time() - t_start
        self.root.after(0, lambda: self.extract_finish(success, fail, skipped, out_dir, duration))

    def run_extract_retry_process(self, retry_tasks):
        t_start = time.time()
        self._process_start_time = t_start
        total = len(retry_tasks)
        if total == 0:
            self.root.after(0, self._update_retry_button_state)
            return

        success = 0
        fail = 0
        skipped = 0
        out_dir = retry_tasks[0].get('out_dir', self.entry_dest.get())
        self.progress['maximum'] = total

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = {}
            planned_names = set()
            skipped_before_submit = 0
            for task in retry_tasks:
                if not self.is_running:
                    self.extract_failed_tasks.append(task)
                    continue
                filename_base = task['filename_base']
                if filename_base in planned_names or self._extract_output_exists(task['out_dir'], filename_base):
                    skipped += 1
                    skipped_before_submit += 1
                    planned_names.add(filename_base)
                    self.root.after(
                        0,
                        self.update_progress_ext,
                        skipped_before_submit,
                        total,
                        success,
                        fail,
                        skipped,
                        self.T['msg_same_name_skip'].format(filename_base)
                    )
                    continue
                planned_names.add(filename_base)
                extract_options = task.get('extract_options')
                if extract_options:
                    future = executor.submit(
                        self.download_url,
                        task['url'],
                        filename_base,
                        task['out_dir'],
                        extract_options
                    )
                else:
                    future = executor.submit(
                        self.download_url,
                        task['url'],
                        filename_base,
                        task['out_dir']
                    )
                futures[future] = task

            completed = skipped_before_submit
            for future in concurrent.futures.as_completed(futures):
                task = futures[future]
                if not self.is_running:
                    self.extract_failed_tasks.append(task)
                    continue
                try:
                    is_ok, msg = future.result()
                except Exception as e:
                    is_ok, msg = False, self.T['msg_err'].format(
                        task['filename_base'],
                        f"{type(e).__name__}: {str(e)[:60]}"
                    )
                completed += 1
                if is_ok:
                    success += 1
                else:
                    fail += 1
                    self.extract_failed_tasks.append(task)
                self.root.after(0, self.update_progress_ext, completed, total, success, fail, skipped, msg)

        duration = time.time() - t_start
        self.root.after(0, lambda: self.extract_finish(success, fail, skipped, out_dir, duration))

    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB

    def _remove_file_quietly(self, path):
        if not path:
            return
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass

    def _write_extract_file_atomic(self, path, data):
        temp_path = path + ".part"
        self._remove_file_quietly(temp_path)
        try:
            with open(temp_path, 'wb') as f:
                f.write(data)
            os.replace(temp_path, path)
            return True
        except OSError:
            self._remove_file_quietly(temp_path)
            return False

    def _save_processed_extract_image(self, image_data, filename_base, out_dir, options):
        try:
            pil_img = PILImage.open(BytesIO(image_data))
            pil_img.load()
            buf, ext = _prepare_extract_image_bytes(
                pil_img,
                bg_mode=options.get('bg_mode', EXTRACT_BG_ORIGINAL),
                shape=options.get('shape', EXTRACT_SHAPE_ORIGINAL),
                add_border=bool(options.get('add_border')),
            )
        except Exception as e:
            return False, self.T['msg_bad_image'].format(filename_base, str(e)[:60])

        path = os.path.join(out_dir, filename_base + ext)
        if not self._write_extract_file_atomic(path, buf.getvalue()):
            return False, self.T['msg_err'].format(filename_base, "Could not write file")
        return True, "OK"

    def _extract_output_exists(self, out_dir, filename_base):
        try:
            for name in os.listdir(out_dir):
                stem, ext = os.path.splitext(name)
                if stem == filename_base and ext:
                    path = os.path.join(out_dir, name)
                    try:
                        if os.path.isfile(path) and os.path.getsize(path) == 0:
                            self._remove_file_quietly(path)
                            continue
                    except OSError:
                        pass
                    return True
        except OSError:
            return False
        return False

    def download_url(self, url, filename_base, out_dir, extract_options=None):
        if not self.is_running:
            return False, "Stopped"
        process_image = _extract_options_require_processing(extract_options)
        for attempt in range(EXTRACT_TIMEOUT_RETRIES + 1):
            path = None
            temp_path = None
            try:
                headers = {'User-Agent': 'Mozilla/5.0'}
                r = requests.get(url, headers=headers, timeout=10, stream=True)
                if not self.is_running:
                    return False, "Stopped"
                if r.status_code == 200:
                    cl = int(r.headers.get('Content-Length', 0))
                    if cl > self.MAX_FILE_SIZE:
                        return False, self.T['msg_too_large'].format(filename_base, cl // 1024 // 1024)
                    ct = r.headers.get('Content-Type', '').lower()
                    ext = mimetypes.guess_extension(ct)
                    if not ext:
                        ext = ".jpg"
                    path = os.path.join(out_dir, filename_base + ext)
                    written = 0
                    if process_image:
                        chunks = []
                        for chunk in r.iter_content(8192):
                            if not self.is_running:
                                return False, "Stopped"
                            written += len(chunk)
                            if written > self.MAX_FILE_SIZE:
                                return False, self.T['msg_too_large'].format(filename_base, written // 1024 // 1024)
                            chunks.append(chunk)
                        if written == 0:
                            if attempt < EXTRACT_TIMEOUT_RETRIES:
                                continue
                            return False, self.T['msg_err'].format(filename_base, "Empty download")
                        return self._save_processed_extract_image(
                            b''.join(chunks),
                            filename_base,
                            out_dir,
                            extract_options
                        )
                    else:
                        temp_path = path + ".part"
                        self._remove_file_quietly(temp_path)
                        with open(temp_path, 'wb') as f:
                            for chunk in r.iter_content(8192):
                                if not self.is_running:
                                    self._remove_file_quietly(temp_path)
                                    return False, "Stopped"
                                written += len(chunk)
                                if written > self.MAX_FILE_SIZE:
                                    f.close()
                                    self._remove_file_quietly(temp_path)
                                    return False, self.T['msg_too_large'].format(filename_base, written // 1024 // 1024)
                                f.write(chunk)
                        if written == 0:
                            self._remove_file_quietly(temp_path)
                            if attempt < EXTRACT_TIMEOUT_RETRIES:
                                continue
                            return False, self.T['msg_err'].format(filename_base, "Empty download")
                        os.replace(temp_path, path)
                        return True, "OK"
                elif r.status_code == 404:
                    return False, self.T['msg_404'].format(filename_base)
                else:
                    return False, self.T['msg_err'].format(filename_base, f"HTTP {r.status_code} ({url[:60]})")
            except requests.exceptions.Timeout:
                self._remove_file_quietly(temp_path)
                if attempt < EXTRACT_TIMEOUT_RETRIES:
                    continue
                return False, self.T['msg_timeout'].format(filename_base)
            except requests.exceptions.SSLError as e:
                self._remove_file_quietly(temp_path)
                return False, self.T['msg_ssl_err'].format(filename_base, str(e)[:80])
            except requests.exceptions.ConnectionError as e:
                self._remove_file_quietly(temp_path)
                return False, self.T['msg_conn_err'].format(filename_base, str(e)[:80])
            except Exception as e:
                self._remove_file_quietly(temp_path)
                if attempt == 0:
                    continue
                return False, self.T['msg_err'].format(filename_base, f"{type(e).__name__}: {str(e)[:60]}")
        return False, self.T['msg_err'].format(filename_base, "Max retries exceeded")

    def _format_eta(self, current, total):
        if current <= 0 or not hasattr(self, '_process_start_time'):
            return ""
        elapsed = time.time() - self._process_start_time
        if elapsed < 1 or current < 2:
            return ""
        remaining = elapsed / current * (total - current)
        if remaining < 60:
            return f"  ETA {int(remaining)}s"
        return f"  ETA {int(remaining//60)}m{int(remaining%60)}s"

    def update_progress_ext(self, current, total, success, fail, skipped, msg):
        if not self.is_running:
            return
        self.progress['value'] = current
        eta = self._format_eta(current, total)
        self.lbl_status.config(text=self.T['status_run'].format(current, total, success, fail, skipped) + eta)
        if "OK" not in msg and "Process" not in msg:
            self.log(msg)

    def extract_finish(self, success, fail, skipped, path, duration):
        if self.is_running:
            self.lbl_status.config(text="Done")
            self.progress['value'] = self.progress['maximum']
            msg = self.T['done_msg'].format(duration, success, fail, skipped, path)
            self.log("-" * 20)
            self.log(msg.replace("\n", " "))
            messagebox.showinfo("Done", msg)
            if success > 0 or self._dir_has_files(path):
                self._open_folder(path)
        else:
            self.lbl_status.config(text="Stopped")
        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')
        self._update_retry_button_state()

    # ==========================================
    # 嵌入图片处理
    # ==========================================

    def clean_url(self, url):
        original = url
        url = re.sub(r'!\d+x\d+', '', url)
        url = re.sub(r'\?imageView2/[^&]*', '', url)
        url = re.sub(r'\?x-oss-process=[^&]*', '', url)
        url = re.sub(r'[?&](width|height|w|h|size|resize|quality|format)=[^&]*', '', url)
        url = re.sub(r'\?\d+$', '', url)
        url = re.sub(r'\?&+', '?', url)
        url = re.sub(r'\?$', '', url)
        if url != original:
            self.root.after(0, lambda: self.log(f"URL clean: {original[:60]}... -> {url[:60]}..."))
        return url

    def download_to_bytesio(self, url, max_dim=None, bg_mode=EMBED_BG_WHITE):
        if not self.is_running:
            return False, "Stopped"
        for attempt in range(2):
            try:
                headers = {'User-Agent': 'Mozilla/5.0'}
                r = requests.get(url, headers=headers, timeout=10)
                if not self.is_running:
                    return False, "Stopped"
                if r.status_code == 200:
                    cl = int(r.headers.get('Content-Length', 0))
                    if cl > self.MAX_FILE_SIZE:
                        return False, self.T['msg_too_large'].format(url[:50], cl // 1024 // 1024)
                    try:
                        pil_img = PILImage.open(BytesIO(r.content))
                    except Exception as e:
                        return False, self.T['msg_bad_image'].format(url[:50], str(e)[:60])
                    return True, _prepare_embed_image_bytes(pil_img, max_dim, bg_mode)
                elif r.status_code == 404:
                    return False, self.T['msg_404'].format(url[:50])
                else:
                    return False, self.T['msg_err'].format(url[:50], f"HTTP {r.status_code} ({url[:60]})")
            except requests.exceptions.Timeout:
                if attempt == 0:
                    continue
                return False, self.T['msg_timeout'].format(url[:50])
            except requests.exceptions.SSLError as e:
                return False, self.T['msg_ssl_err'].format(url[:50], str(e)[:80])
            except requests.exceptions.ConnectionError as e:
                return False, self.T['msg_conn_err'].format(url[:50], str(e)[:80])
            except Exception as e:
                if attempt == 0:
                    continue
                return False, self.T['msg_err'].format(url[:50], f"{type(e).__name__}: {str(e)[:60]}")
        return False, self.T['msg_err'].format(url[:50], "Max retries exceeded")

    def run_embed_process(self):
        t_start = time.time()
        self._process_start_time = t_start
        dest = self.entry_dest.get()
        fname = "Clipboard" if self.file_path == "Clipboard" else os.path.splitext(os.path.basename(self.file_path))[0]

        self.root.after(0, lambda: self.log(self.T['log_embed_start']))

        if self.var_original.get():
            max_dim = None
        else:
            try:
                max_dim = int(self.entry_max_dim.get())
            except ValueError:
                max_dim = 500
        bg_mode = self._get_embed_bg_mode()

        sku_col_idx = self.embed_sku_col_idx
        use_url_library = bool(getattr(self, 'embed_use_url_library', False))
        url_col_idx = self.embed_url_col_idx
        image_anchor_col_idx = sku_col_idx if sku_col_idx is not None else url_col_idx
        if (use_url_library and sku_col_idx is None) or (not use_url_library and url_col_idx is None):
            self.root.after(0, self.embed_error_finish, self.T['msg_no_url'])
            return
        if image_anchor_col_idx is None:
            image_anchor_col_idx = sku_col_idx if use_url_library else url_col_idx
        write_original = self.var_write_original.get()
        self._embed_setup_used_original = write_original
        extra_field_names = self._get_selected_url_library_fields() if use_url_library else []

        header_row_excel = self.header_row + 1
        try:
            if write_original:
                out_file, ws, wb_out, img_header_col, header_row_excel = \
                    self._embed_setup_original(fname, image_anchor_col_idx, extra_field_names)
            else:
                out_file, ws, wb_out, img_header_col, header_row_excel = \
                    self._embed_setup_new(fname, image_anchor_col_idx, header_row_excel, extra_field_names)
        except Exception as e:
            self.root.after(0, self.embed_error_finish, f"{type(e).__name__}: {e}")
            return
        write_original = self._embed_setup_used_original

        orig_cols = list(self.df.columns)
        total = len(self.df)
        self.progress['maximum'] = total

        # Collect URLs
        rows_data = []
        row_library_records = []
        library_matches = 0
        for i in range(total):
            record = {}
            if use_url_library:
                code = _normalize_lookup_code(self.df.iloc[i, sku_col_idx])
                url = getattr(self, 'url_library', {}).get(code)
                record = getattr(self, 'url_library_records', {}).get(code, {})
                if url:
                    library_matches += 1
                    url = self.clean_url(url)
            else:
                url_raw_value = self.df.iloc[i, url_col_idx]
                url_raw = str(url_raw_value).strip()
                url = _extract_http_url(url_raw_value)
                if url:
                    url = self.clean_url(url)
                elif url_raw and url_raw.lower() != 'nan':
                        self.root.after(0, lambda u=url_raw: self.log(
                            self.T['msg_invalid_url'].format(f"Row {i+1}", u[:60])))
            rows_data.append(url)
            row_library_records.append(record)

        if use_url_library:
            self.root.after(0, lambda: self.log(
                self.T['msg_url_lib_matches'].format(library_matches, total)
            ))

        success = 0
        fail = 0
        row_results = [None] * total

        # Download concurrently
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = {}
            for i, url in enumerate(rows_data):
                if not self.is_running:
                    break
                if url:
                    futures[executor.submit(self.download_to_bytesio, url, max_dim, bg_mode)] = i
                else:
                    futures[executor.submit(lambda: (False, "No URL"))] = i

            completed = 0
            for future in concurrent.futures.as_completed(futures):
                if not self.is_running:
                    break
                row_idx = futures[future]
                try:
                    is_ok, result = future.result()
                except Exception as e:
                    is_ok, result = False, str(e)
                row_results[row_idx] = (is_ok, result)
                completed += 1
                self.root.after(0, self.update_progress_emb, completed, total, success, fail)

        # Embed images into sheet
        for i, result in enumerate(row_results):
            if not self.is_running:
                break
            if result is None:
                result = (False, "Stopped")

            is_ok, data = result

            excel_row = header_row_excel + 1 + i
            if not write_original:
                # Write cell values for new-workbook mode
                out_col = 1
                for j in range(len(orig_cols)):
                    cell_val = str(self.df.iloc[i, j]) if self.df.iloc[i, j] is not None else ""
                    if cell_val.lower() == 'nan':
                        cell_val = ""
                    ws.cell(row=excel_row, column=out_col, value=cell_val)
                    out_col += 1
                    if j == image_anchor_col_idx:
                        out_col += 1 + len(extra_field_names)

            img_col_letter = get_column_letter(img_header_col)
            if extra_field_names:
                record = row_library_records[i] if i < len(row_library_records) else {}
                for offset, field_name in enumerate(extra_field_names, start=1):
                    ws.cell(
                        row=excel_row,
                        column=img_header_col + offset,
                        value=_json_safe_value(record.get(field_name, ''))
                    )
            if is_ok:
                try:
                    xl_img = XlImage(data)
                    row_height_pt = 40
                    ws.row_dimensions[excel_row].height = row_height_pt
                    img_ratio = xl_img.width / xl_img.height if xl_img.height > 0 else 1
                    col_width = row_height_pt * 1.33 * img_ratio / 7 + 1
                    ws.column_dimensions[img_col_letter].width = max(col_width, 12)
                    scaled_h = int(row_height_pt * 1.33)
                    scaled_w = int(scaled_h * img_ratio)
                    xl_img.width = scaled_w
                    xl_img.height = scaled_h
                    ws.add_image(xl_img, f"{img_col_letter}{excel_row}")
                    success += 1
                except Exception:
                    ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_fail'])
                    fail += 1
            else:
                ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_fail'])
                fail += 1

        # Handle stopped rows
        if not self.is_running:
            for i in range(total):
                if row_results[i] is None:
                    excel_row = header_row_excel + 1 + i
                    if not write_original:
                        out_col = 1
                        for j in range(len(orig_cols)):
                            cell_val = str(self.df.iloc[i, j]) if self.df.iloc[i, j] is not None else ""
                            if cell_val.lower() == 'nan':
                                cell_val = ""
                            ws.cell(row=excel_row, column=out_col, value=cell_val)
                            out_col += 1
                            if j == image_anchor_col_idx:
                                out_col += 1 + len(extra_field_names)
                    ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_skip'])

        self.root.after(0, lambda: self.log(self.T['log_embed_save']))
        try:
            wb_out.save(out_file)
            wb_out.close()
        except Exception as e:
            try:
                wb_out.close()
            except Exception:
                pass
            self.root.after(0, self.embed_error_finish, f"{type(e).__name__}: {e}")
            return

        duration = time.time() - t_start
        self.root.after(0, lambda: self.embed_finish(success, fail, out_file, duration))

    def _embed_setup_new(self, fname, source_col_idx, header_row_excel=1, extra_field_names=None):
        """Create a new workbook for embedding. Returns (out_file, ws, wb, img_header_col, header_row_excel)."""
        extra_field_names = extra_field_names or []
        dest = self.entry_dest.get()
        out_file = os.path.join(dest, f"{fname}_Embedded.xlsx")
        wb_out = openpyxl.Workbook()
        ws = wb_out.active

        orig_cols = list(self.df.columns)
        out_col = 1
        img_header_col = 1
        for i, col_name in enumerate(orig_cols):
            ws.cell(row=header_row_excel, column=out_col, value=col_name)
            out_col += 1
            if i == source_col_idx:
                ws.cell(row=header_row_excel, column=out_col, value="图片")
                img_header_col = out_col
                out_col += 1
                for field_name in extra_field_names:
                    ws.cell(row=header_row_excel, column=out_col, value=field_name)
                    out_col += 1

        return out_file, ws, wb_out, img_header_col, header_row_excel

    def _embed_setup_original(self, fname, source_col_idx, extra_field_names=None):
        """Load original workbook, insert image column. Returns (out_file, ws, wb, img_header_col, header_row_excel)."""
        extra_field_names = extra_field_names or []
        dest = self.entry_dest.get()
        out_file = os.path.join(dest, f"{fname}_WithImages.xlsx")
        header_row_excel = self.header_row + 1

        if self.file_path == "Clipboard" or not os.path.exists(self.file_path):
            # Clipboard mode: fallback to new workbook
            self._embed_setup_used_original = False
            return self._embed_setup_new(fname, source_col_idx, header_row_excel, extra_field_names)

        ext = os.path.splitext(self.file_path)[1].lower()
        if ext != '.xlsx':
            display_ext = ext or "current file"
            self.root.after(0, lambda e=display_ext: self.log(
                self.T['log_embed_format_fallback'].format(e)))
            self._embed_setup_used_original = False
            return self._embed_setup_new(fname, source_col_idx, header_row_excel, extra_field_names)

        wb_out = openpyxl.load_workbook(self.file_path)
        ws = wb_out.active

        # Find header row and the source/anchor column in the Excel sheet.
        source_col_name = str(self.df.columns[source_col_idx])

        source_excel_col = None
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row_excel, column=col_idx).value
            if cell_val is not None and str(cell_val).strip() == source_col_name:
                source_excel_col = col_idx
                break

        if source_excel_col is None:
            # Fallback: search all rows for the header
            for r in range(1, min(ws.max_row + 1, 20)):
                for c in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=r, column=c).value
                    if cell_val is not None and str(cell_val).strip() == source_col_name:
                        source_excel_col = c
                        header_row_excel = r
                        break
                if source_excel_col:
                    break

        if source_excel_col is None:
            # Last resort: use column index directly
            source_excel_col = source_col_idx + 1

        img_header_col = source_excel_col + 1
        ws.insert_cols(img_header_col, amount=1 + len(extra_field_names))
        ws.cell(row=header_row_excel, column=img_header_col, value="图片")
        for offset, field_name in enumerate(extra_field_names, start=1):
            ws.cell(row=header_row_excel, column=img_header_col + offset, value=field_name)

        return out_file, ws, wb_out, img_header_col, header_row_excel

    def update_progress_emb(self, current, total, success, fail):
        if not self.is_running:
            return
        self.progress['value'] = current
        eta = self._format_eta(current, total)
        self.lbl_status.config(text=self.T['embed_status_run'].format(current, total, success, fail) + eta)

    def embed_error_finish(self, error):
        self.lbl_status.config(text="Error")
        try:
            self.progress.stop()
        except Exception:
            pass
        self.log(self.T['msg_embed_error'].format(error))
        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')

    def embed_finish(self, success, fail, path, duration):
        if self.is_running:
            self.lbl_status.config(text="Done")
            self.progress['value'] = self.progress['maximum']
            msg = self.T['msg_embed_done'].format(duration, success, fail, path)
            self.log("-" * 20)
            self.log(msg.replace("\n", " "))
            if success > 0:
                messagebox.showinfo("Done", msg)
                self._open_folder(os.path.dirname(path))
        else:
            self.lbl_status.config(text="Stopped")
        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')

    # ==========================================
    # 通用工具
    # ==========================================

    def _dir_has_files(self, path):
        try:
            return any(os.path.isfile(os.path.join(path, name)) for name in os.listdir(path))
        except OSError:
            return False

    def _open_folder(self, path):
        try:
            if platform.system() == "Darwin":
                subprocess.run(["open", path], check=False)
            else:
                os.startfile(path)
        except Exception:
            pass

    def check_update(self, auto=False):
        """Check GitHub for latest release. auto=True suppresses log messages."""
        API_URL = "https://api.github.com/repos/youngoris/SheetPic/releases/latest"
        RAW_URL = "https://raw.githubusercontent.com/youngoris/SheetPic/main/sheetpic.py"
        PROXIES = [
            "https://ghfast.top/",
            "https://gh-proxy.com/",
        ]
        RELEASES_URL = GITHUB_URL + "/releases/latest"

        def _fetch(url, timeout=5):
            req = urllib.request.Request(url, headers={"User-Agent": "SheetPic"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", errors="ignore")

        def _do():
            remote_ver = None
            dl_url = RELEASES_URL

            # Source 1: GitHub API (works globally, may be blocked in China)
            try:
                text = _fetch(API_URL)
                data = json.loads(text)
                tag = data.get("tag_name", "").lstrip("v")
                if tag:
                    remote_ver = tuple(int(x) for x in tag.split("."))
                    dl_url = data.get("html_url", RELEASES_URL)
            except Exception:
                pass

            # Source 2: Raw file via proxy (fallback for China)
            if remote_ver is None:
                for proxy in PROXIES:
                    try:
                        text = _fetch(proxy + RAW_URL, timeout=8)
                        m = re.search(r'APP_VERSION\s*=\s*"([^"]+)"', text)
                        if m:
                            remote_ver = tuple(int(x) for x in m.group(1).split("."))
                            break
                    except Exception:
                        continue

            if remote_ver is None:
                if not auto:
                    self.root.after(0, lambda: self.log(self.T['update_check_fail']))
                return

            local_ver = tuple(int(x) for x in APP_VERSION.split("."))
            if remote_ver > local_ver:
                remote_tag = ".".join(str(x) for x in remote_ver)
                self.root.after(0, lambda: self._show_update(remote_tag, dl_url))
            elif not auto:
                self.root.after(0, lambda: self.log(self.T['update_none']))

        threading.Thread(target=_do, daemon=True).start()

    def _show_update(self, version, url):
        self.lbl_update.config(text=self.T['update_available'].format(version))
        self.lbl_update.bind("<Button-1>", lambda e: webbrowser.open(url))

    def on_closing(self):
        self.is_running = False
        if self.wb:
            try:
                self.wb.close()
            except Exception:
                pass
        self.root.destroy()
        os._exit(0)


if __name__ == "__main__":
    root = tk.Tk()
    _set_window_icon(root)
    # Try loading tkdnd for drag-and-drop support
    try:
        root.tk.call('package', 'require', 'tkdnd')
    except tk.TclError:
        pass
    app = SheetPicApp(root)
    root.mainloop()
