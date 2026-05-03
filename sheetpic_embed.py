import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import os
import threading
import platform
import locale
import requests
import concurrent.futures
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XlImage
from PIL import Image as PILImage
from io import BytesIO
import webbrowser
import datetime
import re

# ==========================================
# 语言与配置
# ==========================================
LANG_MAP = {
    'zh': {
        'title': "SheetPic Embed - 图片嵌入Excel助手",
        'footer_text': "SheetPic Embed by Andre",
        'sec_source': "数据来源",
        'sec_settings': "嵌入设置",
        'btn_browse': "📂 选择文件",
        'btn_clip': "📋 剪贴板",
        'lbl_dest': "保存位置:",
        'btn_dest': "修改",
        'lbl_url_col': "图片URL列 (含链接的列)",
        'lbl_sku_col': "SKU/ID列 (用于排序)",
        'lbl_img_size': "最大边长 (px)",
        'chk_original': "插入原图 (不缩放)",
        'btn_start': "开始嵌入",
        'btn_stop': "停止",
        'status_idle': "准备就绪",
        'status_run': "进度: {}/{} (成功: {} | 失败: {})",
        'status_stop': "正在停止...",
        'log_ready': "就绪。请加载含图片URL的表格文件。",
        'log_header': "✅ 锁定表头: 第 {} 行",
        'log_stats': "📊 列分析: 列 {} 含 {} 条有效数据 (类型: URL)",
        'msg_no_url': "❌ 未检测到包含URL的列",
        'msg_embed_done': "耗时: {:.1f}s\n嵌入成功: {}\n下载失败: {}\n输出文件: {}",
        'msg_dl_fail': "[下载失败]",
        'msg_dl_skip': "[无URL]",
        'log_embed_start': "开始嵌入图片处理...",
        'log_embed_save': "正在保存Excel文件...",
        'embed_status_run': "嵌入: {}/{} (成功: {} | 失败: {})",
        'chk_del_url': "嵌入后删除原URL列",
        'msg_404': "❌ {}: [404] 链接失效/不存在",
        'msg_timeout': "⚠️ {}: [超时] 网络连接卡顿",
        'msg_err': "❌ {}: [错误] {}",
    },
    'en': {
        'title': "SheetPic Embed - Image Embedding Tool",
        'footer_text': "SheetPic Embed by Andre",
        'sec_source': "Data Source",
        'sec_settings': "Embed Settings",
        'btn_browse': "📂 File",
        'btn_clip': "📋 Clip",
        'lbl_dest': "Output:",
        'btn_dest': "Change",
        'lbl_url_col': "Image URL Column",
        'lbl_sku_col': "SKU/ID Column (for ordering)",
        'lbl_img_size': "Max Dimension (px)",
        'chk_original': "Original Size (no resize)",
        'btn_start': "Start Embed",
        'btn_stop': "Stop",
        'status_idle': "Ready",
        'status_run': "{} / {} (OK: {} Fail: {})",
        'status_stop': "Stopping...",
        'log_ready': "Ready. Load a table with image URLs.",
        'log_header': "✅ Header at Row {}",
        'log_stats': "📊 Col Stats: {} has {} valid items (URL)",
        'msg_no_url': "❌ No URL column detected",
        'msg_embed_done': "Time: {:.1f}s\nEmbedded: {}\nFailed: {}\nOutput: {}",
        'msg_dl_fail': "[Download Failed]",
        'msg_dl_skip': "[No URL]",
        'log_embed_start': "Starting image embedding...",
        'log_embed_save': "Saving Excel file...",
        'embed_status_run': "Embed: {} / {} (OK: {} | Fail: {})",
        'chk_del_url': "Delete URL column after embedding",
        'msg_404': "❌ {}: [404] Not Found",
        'msg_timeout': "⚠️ {}: [Timeout] Connection failed",
        'msg_err': "❌ {}: [Error] {}",
    }
}

COLORS = {
    'bg': '#F3F4F6', 'card': '#FFFFFF', 'primary': '#2563EB', 'primary_hov': '#1D4ED8',
    'danger': '#DC2626', 'text': '#1F2937', 'text_sub': '#6B7280', 'success': '#10B981',
    'border': '#E5E7EB', 'disabled_bg': '#D1D5DB', 'disabled_fg': '#4B5563'
}

GITHUB_URL = "https://github.com/youngoris/SheetPic"


class SheetPicEmbedApp:
    def __init__(self, root):
        self.root = root
        self.setup_lang()
        self.root.title(self.T['title'])
        self.root.configure(bg=COLORS['bg'])
        if platform.system() == "Darwin":
            self.root.geometry("520x620")
        else:
            self.root.geometry("500x600")

        self.default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        self.file_path = None
        self.df = None
        self.wb = None
        self.ws = None
        self.header_row = 0
        self.embed_url_col_idx = 0
        self.embed_sku_col_idx = 0
        self.embed_url_cols = []
        self.is_running = False

        self.setup_style()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_lang(self):
        self.lang = 'en'
        try:
            if platform.system() == "Darwin":
                # macOS: 从系统偏好读取语言列表
                import subprocess
                result = subprocess.run(
                    ['defaults', 'read', '-g', 'AppleLanguages'],
                    capture_output=True, text=True, timeout=3
                )
                if result.returncode == 0 and 'zh' in result.stdout.lower():
                    self.lang = 'zh'
            else:
                # Windows/Linux: 从环境变量检测
                for var in ('LANG', 'LC_ALL', 'LC_MESSAGES'):
                    val = os.environ.get(var, '')
                    if 'zh' in val.lower():
                        self.lang = 'zh'
                        break
        except:
            pass
        self.T = LANG_MAP[self.lang]

    def setup_style(self):
        style = ttk.Style()
        style.theme_use('clam')

        # macOS: PingFang SC; Windows: Microsoft YaHei UI
        if platform.system() == "Darwin":
            base_font = ("PingFang SC", 13)
        else:
            base_font = ("Microsoft YaHei UI", 9)

        style.configure(".", background=COLORS['card'], foreground=COLORS['text'], font=base_font)
        style.configure("TFrame", background=COLORS['card'])
        style.configure("TEntry", fieldbackground="#F9FAFB", bordercolor=COLORS['border'], padding=5)

        style.configure("TButton", background="#F3F4F6", foreground=COLORS['text'], borderwidth=0, font=base_font)
        style.map("TButton", background=[('active', '#E5E7EB'), ('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])

        style.configure("Primary.TButton", background=COLORS['primary'], foreground="white",
                         font=("PingFang SC" if platform.system() == "Darwin" else "Microsoft YaHei UI",
                               13 if platform.system() == "Darwin" else 10, "bold"), borderwidth=0)
        style.map("Primary.TButton",
                   background=[('active', COLORS['primary_hov']), ('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])

        style.configure("Danger.TButton", background=COLORS['danger'], foreground="white",
                         font=("PingFang SC" if platform.system() == "Darwin" else "Microsoft YaHei UI",
                               13 if platform.system() == "Darwin" else 10, "bold"), borderwidth=0)
        style.map("Danger.TButton", background=[('disabled', COLORS['disabled_bg'])],
                   foreground=[('disabled', COLORS['disabled_fg'])])
        style.configure("Green.Horizontal.TProgressbar", background=COLORS['success'],
                         troughcolor="#E5E7EB", bordercolor=COLORS['card'], thickness=6)

    def setup_ui(self):
        # === card1: 数据来源 ===
        card1 = tk.Frame(self.root, bg=COLORS['card'], padx=15, pady=15)
        card1.pack(fill='x', padx=15, pady=(20, 5))
        tk.Label(card1, text=self.T['sec_source'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 8, "bold")).pack(anchor='w', pady=(0, 5))

        row1 = tk.Frame(card1, bg=COLORS['card'])
        row1.pack(fill='x')
        self.entry_path = ttk.Entry(row1)
        self.entry_path.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=3)
        ttk.Button(row1, text=self.T['btn_browse'], width=10, command=self.select_file).pack(side='left', padx=2)
        ttk.Button(row1, text=self.T['btn_clip'], width=8, command=self.load_clipboard).pack(side='left')

        # 输出目录行 (共享)
        row_dest = tk.Frame(card1, bg=COLORS['card'])
        row_dest.pack(fill='x', pady=(10, 0))
        tk.Label(row_dest, text=self.T['lbl_dest'], bg=COLORS['card'], width=8, anchor='w').pack(side='left')
        self.entry_dest = ttk.Entry(row_dest)
        self.entry_dest.insert(0, self.default_dir)
        self.entry_dest.pack(side='left', fill='x', expand=True, padx=5, ipady=3)
        ttk.Button(row_dest, text=self.T['btn_dest'], width=6, command=self.select_folder).pack(side='left')

        # === card2: 嵌入设置 ===
        card2 = tk.Frame(self.root, bg=COLORS['card'], padx=15, pady=15)
        card2.pack(fill='x', padx=15, pady=10)
        tk.Label(card2, text=self.T['sec_settings'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 8, "bold")).pack(anchor='w', pady=(0, 5))

        # URL列选择
        row_url = tk.Frame(card2, bg=COLORS['card'])
        row_url.pack(fill='x', pady=(0, 8))
        tk.Label(row_url, text=self.T['lbl_url_col'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 8)).pack(anchor='w')
        self.combo_url = ttk.Combobox(row_url, state="disabled")
        self.combo_url.pack(fill='x', pady=(2, 0))

        # SKU列选择
        row_sku = tk.Frame(card2, bg=COLORS['card'])
        row_sku.pack(fill='x', pady=(0, 8))
        tk.Label(row_sku, text=self.T['lbl_sku_col'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 8)).pack(anchor='w')
        self.combo_sku = ttk.Combobox(row_sku, state="disabled")
        self.combo_sku.pack(fill='x', pady=(2, 0))

        # 图片尺寸输入
        row_size = tk.Frame(card2, bg=COLORS['card'])
        row_size.pack(fill='x')
        tk.Label(row_size, text=self.T['lbl_img_size'], bg=COLORS['card'],
                 fg=COLORS['text_sub'], font=("Arial", 8)).pack(anchor='w')
        size_frame = tk.Frame(card2, bg=COLORS['card'])
        size_frame.pack(fill='x', pady=(2, 0))
        self.entry_max_dim = ttk.Entry(size_frame, width=8)
        self.entry_max_dim.pack(side='left', padx=(0, 6))
        self.entry_max_dim.insert(0, "500")

        # 插入原图选项
        self.var_original = tk.BooleanVar(value=False)
        chk_original = tk.Checkbutton(card2, text=self.T['chk_original'],
                                       variable=self.var_original, bg=COLORS['card'],
                                       fg=COLORS['text_sub'], font=("Arial", 8),
                                       activebackground=COLORS['card'],
                                       command=self._toggle_max_dim)
        chk_original.pack(anchor='w', pady=(6, 0))

        # 删除URL列勾选框
        self.var_del_url = tk.BooleanVar(value=False)
        chk_del = tk.Checkbutton(card2, text=self.T['chk_del_url'],
                                  variable=self.var_del_url, bg=COLORS['card'],
                                  fg=COLORS['text_sub'], font=("Arial", 8),
                                  activebackground=COLORS['card'])
        chk_del.pack(anchor='w', pady=(8, 0))

        # === 动作区 ===
        action_frame = tk.Frame(self.root, bg=COLORS['bg'])
        action_frame.pack(fill='x', padx=15, pady=5)
        self.progress = ttk.Progressbar(action_frame, orient="horizontal", mode="determinate",
                                         style="Green.Horizontal.TProgressbar")
        self.progress.pack(fill='x', pady=(0, 5))
        self.lbl_status = tk.Label(action_frame, text="...", bg=COLORS['bg'],
                                    fg=COLORS['text_sub'], font=("Arial", 8))
        self.lbl_status.pack(anchor='e')
        btn_box = tk.Frame(action_frame, bg=COLORS['bg'])
        btn_box.pack(fill='x', pady=5)
        self.btn_run = ttk.Button(btn_box, text=self.T['btn_start'], style="Primary.TButton",
                                   command=self.start_thread, state='disabled')
        self.btn_run.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=5)
        self.btn_stop = ttk.Button(btn_box, text=self.T['btn_stop'], style="Danger.TButton",
                                    command=self.stop_thread, state='disabled')
        self.btn_stop.pack(side='right', fill='x', expand=True, padx=(5, 0), ipady=5)

        # === 日志区 ===
        log_frame = tk.Frame(self.root, bg=COLORS['card'], bd=1, relief="flat")
        log_frame.pack(fill='both', expand=True, padx=15, pady=(5, 5))
        self.log_text = scrolledtext.ScrolledText(log_frame, height=5, font=("Consolas", 8),
                                                   bd=0, highlightthickness=0)
        self.log_text.pack(fill='both', expand=True)
        self.log_text.configure(bg="#FAFAFA", fg="#444", padx=10, pady=10, state='normal')

        # === 页脚 ===
        footer = tk.Frame(self.root, bg=COLORS['bg'])
        footer.pack(side='bottom', fill='x', padx=15, pady=8)
        tk.Label(footer, text=self.T['footer_text'], font=("Arial", 8),
                 bg=COLORS['bg'], fg=COLORS['text_sub']).pack(side='left')
        lbl_link = tk.Label(footer, text="GitHub", font=("Arial", 8),
                             bg=COLORS['bg'], fg=COLORS['primary'], cursor="hand2")
        lbl_link.pack(side='right')
        lbl_link.bind("<Button-1>", lambda e: webbrowser.open(GITHUB_URL))

        self.log(self.T['log_ready'])

    def log(self, msg):
        now = datetime.datetime.now().strftime("[%H:%M:%S]")
        self.log_text.insert(tk.END, f"{now} {msg}\n")
        self.log_text.see(tk.END)

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx;*.xls;*.csv;*.html"), ("All", "*.*")])
        if p:
            self.file_path = p
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, os.path.basename(p))
            threading.Thread(target=self.analyze_data, daemon=True).start()

    def select_folder(self):
        d = filedialog.askdirectory()
        if d:
            self.entry_dest.delete(0, tk.END)
            self.entry_dest.insert(0, d)

    def load_clipboard(self):
        self.log(">>> Reading clipboard...")
        try:
            self.df = pd.read_clipboard()
            if not self.df.empty:
                self.file_path = "Clipboard"
                self.wb = None
                self.entry_path.delete(0, tk.END)
                self.entry_path.insert(0, "Clipboard Data")
                self.process_df()
            else:
                self.log("❌ Clipboard empty")
        except Exception as e:
            self.log(f"❌ Error: {e}")

    def find_robust_header(self, file_path):
        try:
            if os.path.splitext(file_path)[1].lower() == '.csv':
                return 0
            df_raw = pd.read_excel(file_path, header=None, nrows=30)
            row_counts = df_raw.count(axis=1)
            if row_counts.empty:
                return 0
            mode_col_count = row_counts.value_counts().idxmax()
            for idx, count in row_counts.items():
                if count >= mode_col_count:
                    return idx
            return 0
        except:
            return 0

    def analyze_data(self):
        self.progress.config(mode='indeterminate')
        self.progress.start(15)
        self.df = None
        self.wb = None
        self.ws = None
        self.header_row = 0

        try:
            ext = os.path.splitext(self.file_path)[1].lower() if self.file_path != "Clipboard" else ""

            if ext == '.xlsx':
                try:
                    self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
                    self.ws = self.wb.active
                except:
                    pass

            if ext in ['.xlsx', '.xls']:
                self.header_row = self.find_robust_header(self.file_path)
                if self.header_row > 0:
                    self.log(self.T['log_header'].format(self.header_row + 1))

            if ext == '.csv':
                try:
                    self.df = pd.read_csv(self.file_path, encoding='utf-8-sig', on_bad_lines='skip')
                except:
                    self.df = pd.read_csv(self.file_path, encoding='gbk', on_bad_lines='skip')
            elif ext == '.html':
                self.df = pd.read_html(self.file_path)[0]
            else:
                self.df = pd.read_excel(self.file_path, header=self.header_row)

        except Exception as e:
            self.log(f"❌ Error: {e}")

        self.progress.stop()
        self.progress.config(mode='determinate')
        self.progress['value'] = 0
        if self.df is not None and not self.df.empty:
            self.process_df()

    def process_df(self):
        self.df = self.df.astype(str)
        cols = list(self.df.columns)

        # URL计数
        url_counts = {}
        for i, c in enumerate(cols):
            sample_has_url = self.df[c].head(50).str.contains("http", case=False).any()
            if sample_has_url:
                real_count = self.df[c].str.contains("http", case=False, na=False).sum()
                if real_count > 0:
                    url_counts[i] = real_count

        self.embed_url_cols = [{'idx': i, 'count': c} for i, c in url_counts.items()]
        self.embed_url_cols.sort(key=lambda x: x['count'], reverse=True)

        # 构建URL列选项
        url_opts = []
        for item in self.embed_url_cols:
            i = item['idx']
            col_letter = get_column_letter(i + 1)
            display_name = f"{cols[i]} ({col_letter})"
            label = f"{display_name} - {item['count']} URLs"
            self.log(self.T['log_stats'].format(col_letter, item['count'], 'URL'))
            url_opts.append(label)

        # 构建SKU列选项（所有列）
        sku_opts = []
        for i, c in enumerate(cols):
            sku_opts.append(f"{c} ({get_column_letter(i + 1)})")

        self.root.after(0, lambda: self.update_ui_lists(url_opts, sku_opts))

    def update_ui_lists(self, url_opts, sku_opts):
        self.combo_url['values'] = url_opts
        if url_opts:
            self.combo_url.current(0)
            self.embed_url_col_idx = self.embed_url_cols[0]['idx']
        else:
            self.log(self.T['msg_no_url'])

        self.combo_sku['values'] = sku_opts
        best = next((x for x in sku_opts if any(k in x.lower()
                      for k in ["code", "sku", "条码", "货号", "id"])), None)
        if best:
            self.combo_sku.set(best)
        elif sku_opts:
            self.combo_sku.current(0)

        self.combo_url.config(state='readonly')
        self.combo_sku.config(state='readonly')

        if url_opts:
            self.btn_run.config(state='normal')

        self.combo_url.bind('<<ComboboxSelected>>', self._on_url_selected)
        self.combo_sku.bind('<<ComboboxSelected>>', self._on_sku_selected)

    def _get_col_index(self, s):
        match = re.search(r'\(([A-Z]+)\)', s)
        if match:
            return column_index_from_string(match.group(1)) - 1
        return 0

    def _on_url_selected(self, event):
        sel = self.combo_url.get()
        self.embed_url_col_idx = self._get_col_index(sel)

    def _on_sku_selected(self, event):
        sel = self.combo_sku.get()
        self.embed_sku_col_idx = self._get_col_index(sel)

    def start_thread(self):
        self.is_running = True
        self.btn_run.config(state='disabled')
        self.btn_stop.config(state='normal')
        threading.Thread(target=self.run_embed_process, daemon=True).start()

    def stop_thread(self):
        self.is_running = False
        self.log(">>> Stopping...")
        self.btn_stop.config(state='disabled')
        self.lbl_status.config(text=self.T['status_stop'])
        self.progress.stop()

    def _toggle_max_dim(self):
        """切换最大边长输入框的启用/禁用状态"""
        if self.var_original.get():
            self.entry_max_dim.config(state='disabled')
        else:
            self.entry_max_dim.config(state='normal')

    def clean_url(self, url):
        """去除URL中的缩略图参数，确保下载原图。
        支持常见CDN格式:
        - 七牛云: !200x200  /  ?imageView2/0/w/200/h/200
        - 阿里云OSS: ?x-oss-process=image/resize,w_200
        - 通用: ?width=200&height=200  /  ?size=200x200  /  ?w=200&h=200
        - 时间戳: ?134216946960000000
        """
        original = url
        # 1. 去除 !200x200 格式的缩略图参数（七牛云等）
        url = re.sub(r'!\d+x\d+', '', url)
        # 2. 去除 ?imageView2/... 格式（七牛云）
        url = re.sub(r'\?imageView2/[^&]*', '', url)
        # 3. 去除 ?x-oss-process=... 格式（阿里云OSS）
        url = re.sub(r'\?x-oss-process=[^&]*', '', url)
        # 4. 去除 ?width=&height= / ?w=&h= / ?size= 格式
        url = re.sub(r'[?&](width|height|w|h|size|resize|quality|format)=[^&]*', '', url)
        # 5. 去除纯数字时间戳参数（如 ?134216946960000000）
        url = re.sub(r'\?\d+$', '', url)
        # 6. 清理残留的 ? 或 &
        url = re.sub(r'\?&+', '?', url)
        url = re.sub(r'\?$', '', url)
        if url != original:
            self.root.after(0, lambda: self.log(f"🔗 URL清洗: {original[:60]}... → {url[:60]}..."))
        return url

    def download_to_bytesio(self, url, max_dim=None):
        """下载图片，按最大边长缩放为BytesIO。max_dim=None时插入原图。"""
        if not self.is_running:
            return False, "Stopped"
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            r = requests.get(url, headers=headers, timeout=10)
            if not self.is_running:
                return False, "Stopped"
            if r.status_code == 200:
                pil_img = PILImage.open(BytesIO(r.content))
                if max_dim:
                    pil_img.thumbnail((max_dim, max_dim), PILImage.LANCZOS)
                buf = BytesIO()
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    pil_img = pil_img.convert('RGBA')
                    pil_img.save(buf, format='PNG')
                else:
                    pil_img = pil_img.convert('RGB')
                    pil_img.save(buf, format='JPEG', quality=85)
                buf.seek(0)
                return True, buf
            elif r.status_code == 404:
                return False, self.T['msg_404'].format(url[:50])
            else:
                return False, self.T['msg_err'].format(url[:50], f"HTTP {r.status_code}")
        except requests.exceptions.Timeout:
            return False, self.T['msg_timeout'].format(url[:50])
        except Exception as e:
            return False, self.T['msg_err'].format(url[:50], str(e))

    def run_embed_process(self):
        """Embed模式主处理线程"""
        import time
        t_start = time.time()
        dest = self.entry_dest.get()
        fname = "Clipboard" if self.file_path == "Clipboard" else os.path.splitext(os.path.basename(self.file_path))[0]
        out_file = os.path.join(dest, f"{fname}_Embedded.xlsx")

        self.root.after(0, lambda: self.log(self.T['log_embed_start']))

        # 读取配置
        if self.var_original.get():
            max_dim = None
        else:
            try:
                max_dim = int(self.entry_max_dim.get())
            except ValueError:
                max_dim = 500

        url_col_idx = self.embed_url_col_idx
        sku_col_idx = self.embed_sku_col_idx

        # 创建输出工作簿
        wb_out = openpyxl.Workbook()
        ws = wb_out.active

        # === 写入表头 ===
        orig_cols = list(self.df.columns)
        del_url = self.var_del_url.get()

        out_col = 1
        for i, col_name in enumerate(orig_cols):
            if del_url and i == url_col_idx:
                # 跳过URL列，在该位置写"图片"列
                ws.cell(row=1, column=out_col, value="图片")
                img_header_col = out_col
                out_col += 1
            else:
                ws.cell(row=1, column=out_col, value=col_name)
                out_col += 1
                if not del_url and i == url_col_idx:
                    ws.cell(row=1, column=out_col, value="图片")
                    img_header_col = out_col
                    out_col += 1

        # === 预处理URL和SKU ===
        total = len(self.df)
        self.progress['maximum'] = total

        rows_data = []
        for i in range(total):
            code = str(self.df.iloc[i, sku_col_idx]).strip()
            if not code or code.lower() == 'nan':
                code = f"Row_{i+1}"
            url_raw = str(self.df.iloc[i, url_col_idx]).strip()
            url = url_raw
            if url and url.lower() != 'nan' and 'http' in url.lower():
                if not url.startswith("http"):
                    m = re.search(r'(https?://[^\s;]+)', url)
                    if m:
                        url = m.group(1)
                url = self.clean_url(url)
            else:
                url = None
            rows_data.append((code, url))

        # === 并发下载 ===
        success = 0
        fail = 0
        row_results = [None] * total

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            futures = {}
            for i, (code, url) in enumerate(rows_data):
                if not self.is_running:
                    break
                if url:
                    futures[executor.submit(self.download_to_bytesio, url, max_dim)] = i
                else:
                    futures[executor.submit(lambda: (False, "No URL"))] = i

            completed = 0
            for future in concurrent.futures.as_completed(futures):
                if not self.is_running:
                    break
                row_idx = futures[future]
                try:
                    is_ok, result = future.result()
                except Exception as e:
                    is_ok, result = False, str(e)
                row_results[row_idx] = (is_ok, result)
                completed += 1

                self.root.after(0, self.update_progress, completed, total, success, fail)

        # === 按行顺序写入Excel ===
        for i, result in enumerate(row_results):
            if not self.is_running:
                break
            if result is None:
                result = (False, "Stopped")

            is_ok, data = result
            excel_row = i + 2

            # 写入原始数据
            out_col = 1
            for j in range(len(orig_cols)):
                if del_url and j == url_col_idx:
                    # 跳过URL列，该位置留给图片
                    out_col += 1
                else:
                    cell_val = str(self.df.iloc[i, j]) if self.df.iloc[i, j] is not None else ""
                    if cell_val.lower() == 'nan':
                        cell_val = ""
                    ws.cell(row=excel_row, column=out_col, value=cell_val)
                    out_col += 1
                    if not del_url and j == url_col_idx:
                        out_col += 1  # 跳过"图片"列

            # 填充"图片"列
            img_col_letter = get_column_letter(img_header_col)
            if is_ok:
                try:
                    xl_img = XlImage(data)
                    # 行高固定40pt（默认15pt的2.7倍）
                    row_height_pt = 40
                    ws.row_dimensions[excel_row].height = row_height_pt
                    # 按图片原始比例计算列宽
                    img_ratio = xl_img.width / xl_img.height if xl_img.height > 0 else 1
                    # 行高pt → 像素（1pt ≈ 1.33px），列宽字符 ≈ 像素 / 7
                    col_width = row_height_pt * 1.33 * img_ratio / 7 + 1
                    ws.column_dimensions[img_col_letter].width = max(col_width, 12)
                    # 图片尺寸按行高等比缩放
                    scaled_h = int(row_height_pt * 1.33)  # pt → px
                    scaled_w = int(scaled_h * img_ratio)
                    xl_img.width = scaled_w
                    xl_img.height = scaled_h
                    ws.add_image(xl_img, f"{img_col_letter}{excel_row}")
                    success += 1
                except Exception as e:
                    ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_fail'])
                    fail += 1
            else:
                ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_fail'])
                fail += 1

        # 处理被停止时未完成的行
        if not self.is_running:
            for i in range(total):
                if row_results[i] is None:
                    excel_row = i + 2
                    out_col = 1
                    for j in range(len(orig_cols)):
                        if del_url and j == url_col_idx:
                            out_col += 1
                        else:
                            cell_val = str(self.df.iloc[i, j]) if self.df.iloc[i, j] is not None else ""
                            if cell_val.lower() == 'nan':
                                cell_val = ""
                            ws.cell(row=excel_row, column=out_col, value=cell_val)
                            out_col += 1
                            if not del_url and j == url_col_idx:
                                out_col += 1
                    ws.cell(row=excel_row, column=img_header_col, value=self.T['msg_dl_skip'])

        # 保存文件
        self.root.after(0, lambda: self.log(self.T['log_embed_save']))
        wb_out.save(out_file)

        duration = time.time() - t_start
        self.root.after(0, lambda: self.embed_finish(success, fail, out_file, duration))

    def update_progress(self, current, total, success, fail):
        if not self.is_running:
            return
        self.progress['value'] = current
        self.lbl_status.config(text=self.T['status_run'].format(current, total, success, fail))

    def embed_finish(self, success, fail, path, duration):
        """Embed模式完成处理"""
        if self.is_running:
            self.lbl_status.config(text="Done")
            self.progress['value'] = 0
            msg = self.T['msg_embed_done'].format(duration, success, fail, path)
            self.log("-" * 20)
            self.log(msg.replace("\n", " "))
            if success > 0:
                messagebox.showinfo("Done", msg)
                try:
                    if platform.system() == "Darwin":
                        os.system(f'open "{os.path.dirname(path)}"')
                    else:
                        os.startfile(os.path.dirname(path))
                except:
                    pass
        else:
            self.lbl_status.config(text="Stopped")

        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')

    def on_closing(self):
        self.is_running = False
        self.root.destroy()
        os._exit(0)


if __name__ == "__main__":
    root = tk.Tk()
    app = SheetPicEmbedApp(root)
    root.mainloop()
