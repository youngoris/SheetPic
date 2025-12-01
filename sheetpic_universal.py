import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import openpyxl
import os
import threading
import platform # 用于检测操作系统
import subprocess # 用于在 Mac 上打开文件夹
from openpyxl.utils import get_column_letter, column_index_from_string

class SheetPicApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SheetPic - Excel 图片提取神器")
        
        # 根据系统调整窗口大小 (Mac 字体渲染通常较大，稍微宽一点)
        if platform.system() == "Darwin":
            self.root.geometry("720x680")
        else:
            self.root.geometry("700x650")
        
        # 变量初始化
        self.file_path = None
        self.wb = None
        self.ws = None
        
        # === 跨平台：获取桌面路径 ===
        # os.path.expanduser("~") 会自动识别 Mac 的 /Users/xxx 和 Windows 的 C:\Users\xxx
        self.default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        
        self.setup_ui()

    def setup_ui(self):
        # ==========================================
        # 第一步：选择 Excel 文件
        # ==========================================
        frame_step1 = tk.LabelFrame(self.root, text="1. 选择 Excel 文件", padx=10, pady=10, fg="#333333")
        frame_step1.pack(fill='x', padx=15, pady=5)
        
        self.entry_path = tk.Entry(frame_step1, width=50, state='readonly')
        self.entry_path.pack(side='left', padx=5, fill='x', expand=True)
        
        btn_select_file = tk.Button(frame_step1, text="📂 浏览...", command=self.select_file)
        btn_select_file.pack(side='left')

        # ==========================================
        # 第二步：选择导出位置
        # ==========================================
        frame_step2 = tk.LabelFrame(self.root, text="2. 图片保存位置", padx=10, pady=10, fg="#333333")
        frame_step2.pack(fill='x', padx=15, pady=5)
        
        self.entry_dest = tk.Entry(frame_step2, width=50)
        self.entry_dest.insert(0, self.default_dir)
        self.entry_dest.pack(side='left', padx=5, fill='x', expand=True)
        
        btn_select_dest = tk.Button(frame_step2, text="📂 修改...", command=self.select_folder)
        btn_select_dest.pack(side='left')

        # ==========================================
        # 第三步：列识别设置
        # ==========================================
        frame_step3 = tk.LabelFrame(self.root, text="3. 确认列信息 (自动分析)", padx=10, pady=10, fg="#333333")
        frame_step3.pack(fill='x', padx=15, pady=5)
        
        # Grid 布局调整
        frame_step3.columnconfigure(1, weight=1)

        tk.Label(frame_step3, text="图片所在的列:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky='w', pady=5)
        self.combo_img_col = ttk.Combobox(frame_step3, state="disabled")
        self.combo_img_col.grid(row=0, column=1, padx=10, pady=5, sticky='ew')
        
        tk.Label(frame_step3, text="命名依据的列:", font=("Arial", 10, "bold")).grid(row=1, column=0, sticky='w', pady=5)
        self.combo_code_col = ttk.Combobox(frame_step3, state="disabled")
        self.combo_code_col.grid(row=1, column=1, padx=10, pady=5, sticky='ew')

        # ==========================================
        # 底部：按钮与日志
        # ==========================================
        # 注意：Mac 上 Button 的 bg 颜色可能不显示，这是 Tkinter 在 macOS 上的原生限制
        self.btn_run = tk.Button(self.root, text="🚀 开始导出", command=self.start_export_thread, 
                                 font=("Arial", 14, "bold"), height=1, state='disabled',
                                 bg="#4CAF50", fg="black") # Windows 会显示绿色，Mac 显示默认灰
        self.btn_run.pack(fill='x', padx=25, pady=15)
        
        self.log_text = scrolledtext.ScrolledText(self.root, height=12)
        self.log_text.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        self.log(f"当前系统: {platform.system()} {platform.release()}")
        self.log("准备就绪。")

    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
        if path:
            self.file_path = path
            self.entry_path.config(state='normal')
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, path)
            self.entry_path.config(state='readonly')
            threading.Thread(target=self.analyze_file, daemon=True).start()

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.entry_dest.delete(0, tk.END)
            self.entry_dest.insert(0, folder)

    def analyze_file(self):
        self.btn_run.config(state='disabled')
        self.log(f"\n正在分析: {os.path.basename(self.file_path)} ...")
        
        try:
            self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
            self.ws = self.wb.active
            
            # 1. 分析图片列
            img_counts = {}
            images = getattr(self.ws, '_images', [])
            if not images:
                self.log("⚠️ 警告：在该表格中未检测到图片！")
                return

            for img in images:
                try:
                    col_idx = img.anchor._from.col
                    img_counts[col_idx] = img_counts.get(col_idx, 0) + 1
                except: pass
            
            img_options = []
            best_img_col = None
            max_imgs = 0
            for col_idx, count in img_counts.items():
                col_letter = get_column_letter(col_idx + 1)
                opt = f"列 {col_letter} (含 {count} 张图)"
                img_options.append(opt)
                if count > max_imgs:
                    max_imgs = count
                    best_img_col = opt
            
            # 2. 分析条码列
            code_options = []
            best_code_col = None
            scan_limit = min(self.ws.max_column, 26) # 只看前26列(A-Z)
            
            for col in range(1, scan_limit + 1):
                col_letter = get_column_letter(col)
                header = str(self.ws.cell(1, col).value or "无表头")
                sample = str(self.ws.cell(2, col).value or "")[:10]
                
                opt = f"列 {col_letter} - [{header}] (预览: {sample})"
                code_options.append(opt)
                
                # 智能关键词匹配
                keywords = ["条码", "条形码", "编码", "货号", "SKU", "code", "barcode", "id"]
                if any(k in header.lower() for k in keywords):
                    best_code_col = opt
                elif not best_code_col and col_letter == "E":
                    best_code_col = opt

            # 更新 UI
            self.root.after(0, lambda: self.update_ui(img_options, best_img_col, code_options, best_code_col))
            
        except Exception as e:
            self.log(f"❌ 分析失败: {e}")

    def update_ui(self, img_opts, def_img, code_opts, def_code):
        self.combo_img_col['values'] = img_opts
        if def_img: self.combo_img_col.set(def_img)
        elif img_opts: self.combo_img_col.current(0)
        self.combo_img_col.config(state='readonly')
        
        self.combo_code_col['values'] = code_opts
        if def_code: self.combo_code_col.set(def_code)
        elif code_opts: self.combo_code_col.current(0)
        self.combo_code_col.config(state='readonly')
        
        self.btn_run.config(state='normal')
        self.log("✅ 分析完成！请点击开始。")

    def start_export_thread(self):
        dest_dir = self.entry_dest.get()
        img_sel = self.combo_img_col.get()
        code_sel = self.combo_code_col.get()
        
        if not os.path.isdir(dest_dir):
            messagebox.showerror("错误", "保存路径不存在！")
            return
        if not img_sel or not code_sel:
            messagebox.showwarning("提示", "请检查列设置！")
            return
            
        self.btn_run.config(state='disabled')
        threading.Thread(target=self.run_export, args=(dest_dir, img_sel, code_sel)).start()

    def run_export(self, base_dest_dir, img_sel, code_sel):
        try:
            # 创建子文件夹
            file_name = os.path.splitext(os.path.basename(self.file_path))[0]
            final_output_dir = os.path.join(base_dest_dir, f"{file_name}_Images")
            
            if not os.path.exists(final_output_dir):
                os.makedirs(final_output_dir)
            
            self.log(f"\n>>> 文件夹已创建: {final_output_dir}")
            
            # 解析列索引
            target_img_col_letter = img_sel.split(" ")[1]
            target_img_idx = column_index_from_string(target_img_col_letter) - 1
            
            target_code_col_letter = code_sel.split(" ")[1]
            target_code_idx = column_index_from_string(target_code_col_letter)
            
            images = getattr(self.ws, '_images', [])
            count = 0
            
            for i, image in enumerate(images):
                try:
                    if image.anchor._from.col != target_img_idx: continue
                    
                    row = image.anchor._from.row
                    code_val = self.ws.cell(row=row+1, column=target_code_idx).value
                    
                    if code_val:
                        fname = str(code_val).strip()
                        safe_name = "".join([c for c in fname if c.isalnum() or c in ('-','_')]).strip()
                        if not safe_name: safe_name = f"Row_{row+1}"
                        
                        ext = image.format.lower() if image.format else 'jpg'
                        save_path = os.path.join(final_output_dir, f"{safe_name}.{ext}")
                        
                        img_data = None
                        if hasattr(image, '_data'):
                            img_data = image._data() if callable(image._data) else image._data
                        elif hasattr(image, 'ref'):
                            img_data = image.ref.read()
                            
                        if img_data:
                            with open(save_path, "wb") as f:
                                f.write(img_data)
                            self.log(f"导出: {safe_name}.{ext}")
                            count += 1
                except: pass
            
            self.log(f"\n======== 完成 ========")
            self.log(f"成功导出 {count} 张图片。")
            messagebox.showinfo("成功", f"导出完成！\n已保存至: {final_output_dir}")
            
            # === 跨平台：打开文件夹 ===
            self.open_folder_cross_platform(final_output_dir)
            
        except Exception as e:
            self.log(f"错误: {e}")
            messagebox.showerror("出错", str(e))
        finally:
            self.root.after(0, lambda: self.btn_run.config(state='normal'))

    def open_folder_cross_platform(self, path):
        """兼容 Windows 和 macOS 的打开文件夹方法"""
        try:
            current_os = platform.system()
            if current_os == "Windows":
                os.startfile(path)
            elif current_os == "Darwin": # macOS
                subprocess.call(["open", path])
            else: # Linux
                subprocess.call(["xdg-open", path])
        except Exception as e:
            self.log(f"尝试打开文件夹时出错: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SheetPicApp(root)
    root.mainloop()
