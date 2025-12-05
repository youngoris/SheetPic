import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import os
import threading
import platform
import locale
import requests
import concurrent.futures
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import webbrowser
import datetime
import mimetypes

# ==========================================
# 语言与配置
# ==========================================
LANG_MAP = {
    'zh': {
        'title': "SheetPic - 批量图片下载助手", # 保持标题简洁
        'footer_text': "SheetPic by Andre",
        'sec_source': "数据来源",
        'sec_settings': "匹配与保存",
        'btn_browse': "📂 选择文件",
        'btn_clip': "📋 剪贴板",
        'lbl_dest': "保存位置:",
        'btn_dest': "修改",
        'lbl_img': "图片来源 (默认智能合并)",
        'lbl_code': "文件名列 (ID/SKU)",
        'btn_start': "开始处理",
        'btn_stop': "停止",
        'status_idle': "准备就绪",
        'status_run': "进度: {}/{} (成功: {} | 失败: {} | 跳过: {})",
        'status_stop': "正在停止...",
        'log_ready': "已就绪。已启用全量精准计数。",
        'log_header': "✅ 锁定表头: 第 {} 行",
        'log_stats': "📊 列分析: 列 {} 含 {} 条有效数据 (类型: {})",
        'opt_auto': "★ [智能合并] 优先下载数据最多的列 (推荐)",
        'type_url': "[链接] {} (含 {} 个URL)",
        'type_img': "[图片] {} (含 {} 张嵌入图)",
        'msg_skip': "❌ {}: [空] 未检测到有效图片",
        'msg_404': "❌ {}: [404] 链接失效/不存在",
        'msg_timeout': "⚠️ {}: [超时] 网络连接卡顿",
        'msg_err': "❌ {}: [错误] {}",
        'done_msg': "耗时: {:.1f}s\n成功: {}\n失败: {}\n跳过: {}\n保存至: {}"
    },
    'en': {
        'title': "SheetPic - Batch Image Downloader",
        'footer_text': "SheetPic by Andre",
        'sec_source': "Data Source",
        'sec_settings': "Settings",
        'btn_browse': "📂 File",
        'btn_clip': "📋 Clip",
        'lbl_dest': "Output:",
        'btn_dest': "Change",
        'lbl_img': "Image Source (Auto Merge)",
        'lbl_code': "Filename Column",
        'btn_start': "Start",
        'btn_stop': "Stop",
        'status_idle': "Ready",
        'status_run': "{} / {} (OK: {} Fail: {} Skip: {})",
        'status_stop': "Stopping...",
        'log_ready': "Ready. Precise counting enabled.",
        'log_header': "✅ Header at Row {}",
        'log_stats': "📊 Col Stats: {} has {} valid items ({})",
        'opt_auto': "★ [Auto Merge] Priority by count",
        'type_url': "[Link] {} ({} URLs)",
        'type_img': "[Image] {} ({} Embedded)",
        'msg_skip': "❌ {}: [Skip] No valid image found",
        'msg_404': "❌ {}: [404] Not Found",
        'msg_timeout': "⚠️ {}: [Timeout] Connection failed",
        'msg_err': "❌ {}: [Error] {}",
        'done_msg': "Time: {:.1f}s\nSuccess: {}\nFailed: {}\nSkipped: {}\nPath: {}"
    }
}

COLORS = {
    'bg': '#F3F4F6', 'card': '#FFFFFF', 'primary': '#2563EB', 'primary_hov': '#1D4ED8',
    'danger': '#DC2626', 'text': '#1F2937', 'text_sub': '#6B7280', 'success': '#10B981',
    'border': '#E5E7EB', 'disabled_bg': '#D1D5DB', 'disabled_fg': '#4B5563'
}

GITHUB_URL = "https://github.com/youngoris/SheetPic"

class SheetPicApp:
    def __init__(self, root):
        self.root = root
        self.setup_lang()
        self.root.title(self.T['title'])
        self.root.configure(bg=COLORS['bg'])
        if platform.system() == "Darwin": self.root.geometry("520x600")
        else: self.root.geometry("500x580")
        
        self.default_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        self.file_path = None
        self.df = None
        self.wb = None
        self.ws = None
        self.header_row = 0
        self.sorted_img_cols = [] 
        self.is_running = False
        
        self.setup_style()
        self.setup_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def setup_lang(self):
        self.lang = 'en'
        try:
            sys_lang = locale.getdefaultlocale()[0]
            if sys_lang and 'zh' in sys_lang.lower(): self.lang = 'zh'
        except: pass
        self.T = LANG_MAP[self.lang]

    def setup_style(self):
        style = ttk.Style()
        style.theme_use('clam')
        base_font = ("Microsoft YaHei UI", 9)
        style.configure(".", background=COLORS['card'], foreground=COLORS['text'], font=base_font)
        style.configure("TFrame", background=COLORS['card'])
        style.configure("TEntry", fieldbackground="#F9FAFB", bordercolor=COLORS['border'], padding=5)
        
        style.configure("TButton", background="#F3F4F6", foreground=COLORS['text'], borderwidth=0, font=base_font)
        style.map("TButton", background=[('active', '#E5E7EB'), ('disabled', COLORS['disabled_bg'])], foreground=[('disabled', COLORS['disabled_fg'])])
        
        style.configure("Primary.TButton", background=COLORS['primary'], foreground="white", font=("Microsoft YaHei UI", 10, "bold"), borderwidth=0)
        style.map("Primary.TButton", background=[('active', COLORS['primary_hov']), ('disabled', COLORS['disabled_bg'])], foreground=[('disabled', COLORS['disabled_fg'])])
        
        style.configure("Danger.TButton", background=COLORS['danger'], foreground="white", font=("Microsoft YaHei UI", 10, "bold"), borderwidth=0)
        style.map("Danger.TButton", background=[('disabled', COLORS['disabled_bg'])], foreground=[('disabled', COLORS['disabled_fg'])])
        style.configure("Green.Horizontal.TProgressbar", background=COLORS['success'], troughcolor="#E5E7EB", bordercolor=COLORS['card'], thickness=6)

    def setup_ui(self):
        card1 = tk.Frame(self.root, bg=COLORS['card'], padx=15, pady=15)
        card1.pack(fill='x', padx=15, pady=(20, 5))
        tk.Label(card1, text=self.T['sec_source'], bg=COLORS['card'], fg=COLORS['text_sub'], font=("Arial", 8, "bold")).pack(anchor='w', pady=(0, 5))
        row1 = tk.Frame(card1, bg=COLORS['card'])
        row1.pack(fill='x')
        self.entry_path = ttk.Entry(row1)
        self.entry_path.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=3)
        ttk.Button(row1, text=self.T['btn_browse'], width=10, command=self.select_file).pack(side='left', padx=2)
        ttk.Button(row1, text=self.T['btn_clip'], width=8, command=self.load_clipboard).pack(side='left')

        card2 = tk.Frame(self.root, bg=COLORS['card'], padx=15, pady=15)
        card2.pack(fill='x', padx=15, pady=10)
        tk.Label(card2, text=self.T['sec_settings'], bg=COLORS['card'], fg=COLORS['text_sub'], font=("Arial", 8, "bold")).pack(anchor='w', pady=(0, 5))
        row_dest = tk.Frame(card2, bg=COLORS['card'])
        row_dest.pack(fill='x', pady=(0, 10))
        tk.Label(row_dest, text=self.T['lbl_dest'], bg=COLORS['card'], width=8, anchor='w').pack(side='left')
        self.entry_dest = ttk.Entry(row_dest)
        self.entry_dest.insert(0, self.default_dir)
        self.entry_dest.pack(side='left', fill='x', expand=True, padx=5, ipady=3)
        ttk.Button(row_dest, text=self.T['btn_dest'], width=6, command=self.select_folder).pack(side='left')

        row_cols = tk.Frame(card2, bg=COLORS['card'])
        row_cols.pack(fill='x')
        col_box1 = tk.Frame(row_cols, bg=COLORS['card'])
        col_box1.pack(side='left', fill='x', expand=True, padx=(0, 5))
        tk.Label(col_box1, text=self.T['lbl_img'], bg=COLORS['card'], fg=COLORS['text_sub'], font=("Arial", 8)).pack(anchor='w')
        self.combo_img = ttk.Combobox(col_box1, state="disabled")
        self.combo_img.pack(fill='x', pady=(2, 0))
        col_box2 = tk.Frame(row_cols, bg=COLORS['card'])
        col_box2.pack(side='left', fill='x', expand=True, padx=(5, 0))
        tk.Label(col_box2, text=self.T['lbl_code'], bg=COLORS['card'], fg=COLORS['text_sub'], font=("Arial", 8)).pack(anchor='w')
        self.combo_code = ttk.Combobox(col_box2, state="disabled")
        self.combo_code.pack(fill='x', pady=(2, 0))

        action_frame = tk.Frame(self.root, bg=COLORS['bg'])
        action_frame.pack(fill='x', padx=15, pady=5)
        self.progress = ttk.Progressbar(action_frame, orient="horizontal", mode="determinate", style="Green.Horizontal.TProgressbar")
        self.progress.pack(fill='x', pady=(0, 5))
        self.lbl_status = tk.Label(action_frame, text="...", bg=COLORS['bg'], fg=COLORS['text_sub'], font=("Arial", 8))
        self.lbl_status.pack(anchor='e')
        btn_box = tk.Frame(action_frame, bg=COLORS['bg'])
        btn_box.pack(fill='x', pady=5)
        self.btn_run = ttk.Button(btn_box, text=self.T['btn_start'], style="Primary.TButton", command=self.start_thread, state='disabled')
        self.btn_run.pack(side='left', fill='x', expand=True, padx=(0, 5), ipady=5)
        self.btn_stop = ttk.Button(btn_box, text=self.T['btn_stop'], style="Danger.TButton", command=self.stop_thread, state='disabled')
        self.btn_stop.pack(side='right', fill='x', expand=True, padx=(5, 0), ipady=5)

        log_frame = tk.Frame(self.root, bg=COLORS['card'], bd=1, relief="flat")
        log_frame.pack(fill='both', expand=True, padx=15, pady=(5, 5))
        self.log_text = scrolledtext.ScrolledText(log_frame, height=5, font=("Consolas", 8), bd=0, highlightthickness=0)
        self.log_text.pack(fill='both', expand=True)
        self.log_text.configure(bg="#FAFAFA", fg="#444", padx=10, pady=10, state='normal')
        
        footer = tk.Frame(self.root, bg=COLORS['bg'])
        footer.pack(side='bottom', fill='x', padx=15, pady=8)
        tk.Label(footer, text=self.T['footer_text'], font=("Arial", 8), bg=COLORS['bg'], fg=COLORS['text_sub']).pack(side='left')
        lbl_link = tk.Label(footer, text="GitHub 🔗", font=("Arial", 8), bg=COLORS['bg'], fg=COLORS['primary'], cursor="hand2")
        lbl_link.pack(side='right')
        lbl_link.bind("<Button-1>", lambda e: webbrowser.open(GITHUB_URL))

        self.log(self.T['log_ready'])

    def log(self, msg):
        now = datetime.datetime.now().strftime("[%H:%M:%S]")
        self.log_text.insert(tk.END, f"{now} {msg}\n")
        self.log_text.see(tk.END)

    def select_file(self):
        p = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx;*.xls;*.csv;*.html"), ("All", "*.*")])
        if p:
            self.file_path = p
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, os.path.basename(p))
            threading.Thread(target=self.analyze_data, daemon=True).start()

    def select_folder(self):
        d = filedialog.askdirectory()
        if d:
            self.entry_dest.delete(0, tk.END)
            self.entry_dest.insert(0, d)

    def load_clipboard(self):
        self.log(">>> Reading clipboard...")
        try:
            self.df = pd.read_clipboard()
            if not self.df.empty:
                self.file_path = "Clipboard"
                self.wb = None
                self.entry_path.delete(0, tk.END)
                self.entry_path.insert(0, "Clipboard Data")
                self.process_df()
            else: self.log("❌ Clipboard empty")
        except Exception as e: self.log(f"❌ Error: {e}")

    def find_robust_header(self, file_path):
        try:
            if os.path.splitext(file_path)[1].lower() == '.csv': return 0
            df_raw = pd.read_excel(file_path, header=None, nrows=30)
            row_counts = df_raw.count(axis=1)
            if row_counts.empty: return 0
            mode_col_count = row_counts.value_counts().idxmax()
            for idx, count in row_counts.items():
                if count >= mode_col_count: return idx
            return 0
        except: return 0

    def analyze_data(self):
        self.progress.config(mode='indeterminate')
        self.progress.start(15)
        self.df = None
        self.wb = None
        self.header_row = 0
        
        try:
            ext = os.path.splitext(self.file_path)[1].lower() if self.file_path != "Clipboard" else ""
            if ext == '.xlsx':
                try:
                    self.wb = openpyxl.load_workbook(self.file_path, data_only=True)
                    self.ws = self.wb.active
                except: pass

            if ext in ['.xlsx', '.xls']:
                self.header_row = self.find_robust_header(self.file_path)
                if self.header_row > 0: self.log(self.T['log_header'].format(self.header_row + 1))

            if ext == '.csv':
                try: self.df = pd.read_csv(self.file_path, encoding='utf-8-sig', on_bad_lines='skip')
                except: self.df = pd.read_csv(self.file_path, encoding='gbk', on_bad_lines='skip')
            elif ext == '.html': self.df = pd.read_html(self.file_path)[0]
            else: self.df = pd.read_excel(self.file_path, header=self.header_row)
            
        except Exception as e: self.log(f"❌ Error: {e}")
            
        self.progress.stop()
        self.progress.config(mode='determinate')
        self.progress['value'] = 0
        if self.df is not None and not self.df.empty: self.process_df()

    def process_df(self):
        self.df = self.df.astype(str)
        cols = list(self.df.columns)
        img_opts = []
        code_opts = []
        self.sorted_img_cols = []
        
        embed_counts = {}
        if self.wb:
            for img in getattr(self.ws, '_images', []):
                try:
                    c = img.anchor._from.col
                    embed_counts[c] = embed_counts.get(c, 0) + 1
                except: pass

        # === 核心修改：全量 URL 计数 ===
        url_counts = {}
        for i, c in enumerate(cols):
            # 1. 快速初筛：先看前50行有没有 http，如果没有就直接跳过（为了性能）
            sample_has_url = self.df[c].head(50).str.contains("http", case=False).any()
            
            if sample_has_url:
                # 2. 精准计数：如果有希望，则全量扫描该列（Pandas 速度很快，不必担心）
                # 统计包含 "http" 且不是 NaN/空 的单元格数量
                real_count = self.df[c].str.contains("http", case=False, na=False).sum()
                if real_count > 0:
                    url_counts[i] = real_count

        all_img_indices = set(embed_counts.keys()) | set(url_counts.keys())
        for idx in all_img_indices:
            count = max(embed_counts.get(idx, 0), url_counts.get(idx, 0))
            type_str = "embed" if idx in embed_counts else "url"
            self.sorted_img_cols.append({'idx': idx, 'count': count, 'type': type_str})
        
        self.sorted_img_cols.sort(key=lambda x: x['count'], reverse=True)

        if self.sorted_img_cols: img_opts.append(self.T['opt_auto'])
            
        for item in self.sorted_img_cols:
            i = item['idx']
            col_letter = get_column_letter(i + 1)
            display_name = f"{cols[i]} ({col_letter})"
            if item['type'] == 'embed':
                label = self.T['type_img'].format(display_name, item['count'])
                self.log(self.T['log_stats'].format(col_letter, item['count'], "Embedded"))
            else:
                label = self.T['type_url'].format(display_name, item['count'])
                self.log(self.T['log_stats'].format(col_letter, item['count'], "URL"))
            img_opts.append(label)
            
        for i, c in enumerate(cols):
            code_opts.append(f"{c} ({get_column_letter(i+1)})")

        self.root.after(0, lambda: self.update_ui_lists(img_opts, code_opts))

    def update_ui_lists(self, img_opts, code_opts):
        self.combo_img['values'] = img_opts
        if img_opts: self.combo_img.current(0)
        self.combo_code['values'] = code_opts
        best = next((x for x in code_opts if any(k in x.lower() for k in ["code", "sku", "条码", "货号"])), None)
        if best: self.combo_code.set(best)
        elif code_opts: self.combo_code.current(0)
        self.combo_img.config(state='readonly')
        self.combo_code.config(state='readonly')
        if img_opts: self.btn_run.config(state='normal')

    def start_thread(self):
        self.is_running = True
        self.btn_run.config(state='disabled')
        self.btn_stop.config(state='normal')
        threading.Thread(target=self.run_process, daemon=True).start()

    def stop_thread(self):
        self.is_running = False
        self.log(">>> Stopping...")
        self.btn_stop.config(state='disabled')
        self.lbl_status.config(text=self.T['status_stop'])
        self.progress.stop()

    def run_process(self):
        import time
        t_start = time.time()
        dest = self.entry_dest.get()
        fname = "Clipboard" if self.file_path == "Clipboard" else os.path.splitext(os.path.basename(self.file_path))[0]
        out_dir = os.path.join(dest, f"{fname}_Img")
        if not os.path.exists(out_dir): os.makedirs(out_dir)
        
        def get_col_index(s):
            import re
            match = re.search(r'\(([A-Z]+)\)', s)
            if match: return column_index_from_string(match.group(1)) - 1
            return 0
        
        idx_code = get_col_index(self.combo_code.get())
        selection = self.combo_img.get()
        target_cols = []
        
        if "★" in selection: target_cols = self.sorted_img_cols
        else:
            sel_idx = get_col_index(selection)
            for item in self.sorted_img_cols:
                if item['idx'] == sel_idx:
                    target_cols = [item]
                    break
        
        img_map_row_col = {}
        if self.wb:
            for img in getattr(self.ws, '_images', []):
                r = img.anchor._from.row
                c = img.anchor._from.col
                if r not in img_map_row_col: img_map_row_col[r] = {}
                img_map_row_col[r][c] = img

        success = 0
        fail = 0
        skipped = 0
        self.progress['maximum'] = len(self.df)
        tasks = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            for i in range(len(self.df)):
                if not self.is_running: break
                
                code = str(self.df.iloc[i, idx_code]).strip()
                base_name = "".join([c for c in code if c.isalnum() or c in '-_'])
                if not base_name: base_name = f"Row_{i+1}"

                row_images = []
                for col_info in target_cols:
                    c_idx = col_info['idx']
                    if col_info['type'] == 'embed':
                        excel_row = self.header_row + 1 + i
                        if excel_row in img_map_row_col and c_idx in img_map_row_col[excel_row]:
                            img_obj = img_map_row_col[excel_row][c_idx]
                            row_images.append(('embed', img_obj))
                    elif col_info['type'] == 'url':
                        val = str(self.df.iloc[i, c_idx]).strip()
                        # === 空值/无效值 严格过滤 ===
                        if not val or val.lower() == 'nan' or "http" not in val.lower():
                            continue
                            
                        if not val.startswith("http"):
                            import re
                            m = re.search(r'(https?://[^\s;]+)', val)
                            if m: val = m.group(1)
                        val = val.split('?')[0].split('!')[0]
                        row_images.append(('url', val))

                if not row_images:
                    skipped += 1
                    self.root.after(0, self.update_progress, i+1+len(tasks), len(self.df), success, fail, skipped, self.T['msg_skip'].format(base_name))
                    continue

                for img_idx, (src_type, src_data) in enumerate(row_images):
                    suffix = f"-{img_idx}" if img_idx > 0 else ""
                    final_name = f"{base_name}{suffix}"
                    
                    if src_type == 'embed':
                        try:
                            ext = ".png" if src_data.format == "png" else ".jpg"
                            path = os.path.join(out_dir, final_name + ext)
                            with open(path, "wb") as f: f.write(src_data._data())
                            success += 1
                        except: fail += 1
                    else:
                        tasks.append(executor.submit(self.download_url, src_data, final_name, out_dir))
                
                self.root.after(0, self.update_progress, i+1, len(self.df), success, fail, skipped, "Process")

            done_count = i if 'i' in locals() else 0
            for future in concurrent.futures.as_completed(tasks):
                if not self.is_running: break
                is_ok, msg = future.result()
                if is_ok: success += 1
                else: fail += 1
                self.root.after(0, self.update_progress, len(self.df), len(self.df), success, fail, skipped, msg)

        duration = time.time() - t_start
        self.root.after(0, lambda: self.finish(success, fail, skipped, out_dir, duration))

    def download_url(self, url, filename_base, out_dir):
        if not self.is_running: return False, "Stopped"
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            r = requests.get(url, headers=headers, timeout=10, stream=True)
            if not self.is_running: return False, "Stopped"
            if r.status_code == 200:
                ct = r.headers.get('Content-Type', '').lower()
                ext = mimetypes.guess_extension(ct)
                if not ext: ext = ".jpg"
                path = os.path.join(out_dir, filename_base + ext)
                with open(path, 'wb') as f:
                    for chunk in r.iter_content(8192):
                        if not self.is_running: return False, "Stopped"
                        f.write(chunk)
                return True, "OK"
            elif r.status_code == 404: return False, self.T['msg_404'].format(filename_base)
            else: return False, self.T['msg_err'].format(filename_base, f"HTTP {r.status_code}")
        except requests.exceptions.Timeout: return False, self.T['msg_timeout'].format(filename_base)
        except Exception as e: return False, self.T['msg_err'].format(filename_base, str(e))

    def update_progress(self, current, total, success, fail, skipped, msg):
        if not self.is_running: return
        self.progress['value'] = current
        self.lbl_status.config(text=self.T['status_run'].format(current, total, success, fail, skipped))
        if "OK" not in msg and "Process" not in msg:
            self.log(msg)

    def finish(self, success, fail, skipped, path, duration):
        if self.is_running:
            self.lbl_status.config(text="Done")
            self.progress['value'] = 0
            msg = self.T['done_msg'].format(duration, success, fail, skipped, path)
            self.log("-" * 20)
            self.log(msg.replace("\n", " "))
            if success > 0:
                messagebox.showinfo("Done", msg)
                try: os.startfile(path)
                except: pass
        else:
            self.lbl_status.config(text="Stopped")
        
        self.is_running = False
        self.btn_run.config(state='normal')
        self.btn_stop.config(state='disabled')

    def on_closing(self):
        self.is_running = False
        self.root.destroy()
        os._exit(0)

if __name__ == "__main__":
    root = tk.Tk()
    app = SheetPicApp(root)
    root.mainloop()
